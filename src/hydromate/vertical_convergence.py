"""Vertical-layer convergence study for the TELEMAC-3D case.

The 2D study (:mod:`hydromate.convergence`) varies the *horizontal* cell size; this
one varies the **number of sigma layers** (``NUMBER OF HORIZONTAL LEVELS``) on the
*same* horizontal mesh, to demonstrate that the depth-resolved solution is
independent of the vertical discretisation. For a sequence of layer counts
(coarse = few layers -> fine = many) it runs the hotstarted non-hydrostatic 3D case
(:func:`hydromate.threed.build_3d_cas`), samples the depth-averaged quantities of
interest (water depth, scalar velocity, from the 3D run's ``2D RESULT FILE``) at the
ground-truth probe points, and reports the relative change between successive layer
counts plus an observed order of convergence and a Grid Convergence Index.

The "grid size" for the GCI is the layer thickness ``dz`` (which halves as the layer
count roughly doubles), so the same Richardson/GCI maths as the 2D study apply - the
report helpers are reused from :mod:`hydromate.convergence`. The TELEMAC execution is
injected (the ``simulate`` callable), so the report maths are unit-testable without
the solver.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np

from hydromate import threed
from hydromate.core import selafin
from hydromate.config import Config
from hydromate.convergence import (
    DEFAULT_QUANTITIES, _gci_triplet, _rel_change, default_probes,
    extract_at_points,
)

log = logging.getLogger("hydromate")


# --------------------------------------------------------------------------- #
# data model
# --------------------------------------------------------------------------- #

@dataclass
class VerticalLevel:
    """One vertical resolution: its layer count, thickness and sampled QoI."""

    n_levels: int
    dz: float                              # representative layer thickness [m]
    n_nodes3d: int                         # 3D prism mesh node count
    n_elem3d: int                          # 3D prism element count
    time_step: float | None                # fixed dt sized for the target Courant [s]
    runtime_s: float | None
    qoi: dict[str, np.ndarray]             # {quantity: (n_probes,) values}

    @property
    def cell_size(self) -> float:
        """The grid size used for the GCI maths: the layer thickness."""
        return self.dz

    @property
    def label(self) -> str:
        return f"{self.n_levels} levels"

    def repr_value(self, qty: str) -> float:
        return float(np.mean(np.abs(self.qoi[qty])))


@dataclass
class VerticalConvergenceReport:
    levels: list[VerticalLevel]            # ordered coarse (few) -> fine (many)
    quantities: list[str]
    tolerance: float
    probes: np.ndarray
    rel_change: dict[str, list[float]] = field(default_factory=dict)
    observed_order: dict[str, float | None] = field(default_factory=dict)
    gci_fine: dict[str, float | None] = field(default_factory=dict)
    converged: bool = False
    case: str = ""

    def format(self) -> str:
        lv = self.levels
        w = max(12, *(len(x.label) for x in lv))
        lines = [
            "Vertical-layer convergence study (few -> many layers)",
            f"  probes: {len(self.probes)}   tolerance: {self.tolerance * 100:.1f}%",
            f"  {'vertical levels':<22}" + "".join(f"{x.n_levels:>{w}d}" for x in lv),
            f"  {'layer thickness dz [m]':<22}" + "".join(f"{x.dz:>{w}.3f}" for x in lv),
            f"  {'3D nodes':<22}" + "".join(f"{x.n_nodes3d:>{w}d}" for x in lv),
            f"  {'3D elements':<22}" + "".join(f"{x.n_elem3d:>{w}d}" for x in lv),
            f"  {'time step dt [s]':<22}"
            + "".join((f"{x.time_step:>{w}.4f}" if x.time_step else f"{'-':>{w}}")
                      for x in lv),
            f"  {'runtime [s]':<22}"
            + "".join((f"{x.runtime_s:>{w}.1f}" if x.runtime_s else f"{'-':>{w}}")
                      for x in lv),
            "  " + "-" * (22 + w * len(lv)),
            f"  {'quantity phi':<22}" + "".join(f"{x.label:>{w}}" for x in lv),
        ]
        for q in self.quantities:
            lines.append(f"  {q:<22}" + "".join(f"{x.repr_value(q):>{w}.4f}" for x in lv))
            rc = self.rel_change.get(q, [])
            cells = [f"{'':>{w}}"] + [f"{e * 100:>{w - 1}.2f}%" for e in rc]
            lines.append(f"    {'rel. change':<20}" + "".join(cells))
            p, g = self.observed_order.get(q), self.gci_fine.get(q)
            extra = []
            if p is not None:
                extra.append(f"observed order p={p:.2f}")
            if g is not None:
                extra.append(f"GCI(fine)={g * 100:.2f}%")
            if extra:
                lines.append(f"    {'':<20}" + "; ".join(extra))
        verdict = "CONVERGED" if self.converged else "NOT converged"
        lines.append(f"  => {verdict} at {self.tolerance * 100:.1f}% "
                     "(finest relative change vs. tolerance)")
        if not self.converged:
            lines.append("     add layers (the finest still changes the QoI by more "
                         "than the tolerance).")
        return "\n".join(lines)

    def save(self, path) -> Path:
        path = Path(path)
        path.write_text(self.format() + "\n")
        return path

    def recommendation(self) -> dict:
        """Recommend the **fewest** layers that are grid-independent (cheapest 3D run).

        A vertical resolution is adequate when adding more layers changes every
        quantity of interest by less than the tolerance; the coarsest such level is
        the cheapest converged choice."""
        lv = self.levels
        chosen = None
        for i in range(len(lv) - 1):                       # few -> many
            if all(self.rel_change[q][i] < self.tolerance for q in self.quantities):
                chosen = lv[i]
                break
        if chosen is None:
            chosen, converged = lv[-1], False
            reason = (f"no layer count met the {self.tolerance * 100:.0f}% tolerance; "
                      "using the finest (add more layers to confirm independence)")
        else:
            converged = True
            reason = (f"fewest layers whose depth-averaged depth and velocity stay "
                      f"within {self.tolerance * 100:.0f}% under vertical refinement "
                      "(grid-independent and cheapest to run)")
        finest = lv[-1]
        speedup = (finest.runtime_s / chosen.runtime_s
                   if chosen.runtime_s and finest.runtime_s else None)
        return {"level": chosen, "n_levels": chosen.n_levels, "dz": chosen.dz,
                "converged": converged, "reason": reason,
                "speedup_vs_finest": speedup}

    def to_xlsx(self, path) -> Path:
        """Write a styled .xlsx vertical-convergence report with the recommendation."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        path = Path(path)
        lv = self.levels
        head_fill = PatternFill("solid", fgColor="1F4E78")
        head_font = Font(bold=True, color="FFFFFF")
        label_font = Font(bold=True)
        rec_fill = PatternFill("solid", fgColor="C6EFCE")
        warn_fill = PatternFill("solid", fgColor="FFC7CE")
        sub_fill = PatternFill("solid", fgColor="DDEBF7")
        center = Alignment(horizontal="center")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = "vertical convergence"
        ncol = 1 + len(lv)

        def row(cells, *, fill=None, font=None, fmt=None, align=None):
            ws.append(list(cells))
            r = ws.max_row
            for c in range(1, len(cells) + 1):
                cell = ws.cell(row=r, column=c)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if fmt and c > 1 and isinstance(cell.value, (int, float)):
                    cell.number_format = fmt
                if align and c > 1:
                    cell.alignment = align
                cell.border = border
            return r

        title = "Vertical-layer convergence study"
        if self.case:
            title += f" - case '{self.case}'"
        ws.append([title])
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ncol)
        ws["A1"].font = Font(bold=True, size=14)
        ws.append([f"probes: {len(self.probes)}    tolerance: {self.tolerance * 100:.1f}%"
                   f"    layer counts: {len(lv)} (few to many)"])
        ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=ncol)
        ws.append([])

        row(["", *[x.label for x in lv]], fill=head_fill, font=head_font, align=center)
        row(["vertical levels", *[x.n_levels for x in lv]], font=label_font,
            fmt="0", align=center)
        row(["layer thickness dz [m]", *[x.dz for x in lv]], font=label_font,
            fmt="0.000", align=center)
        row(["3D nodes", *[x.n_nodes3d for x in lv]], font=label_font, fmt="#,##0",
            align=center)
        row(["3D elements", *[x.n_elem3d for x in lv]], font=label_font, fmt="#,##0",
            align=center)
        row(["time step dt [s]", *[(x.time_step or float('nan')) for x in lv]],
            font=label_font, fmt="0.0000", align=center)
        row(["runtime [s]", *[(x.runtime_s or float('nan')) for x in lv]],
            font=label_font, fmt="0.0", align=center)
        ws.append([])

        for q in self.quantities:
            row([q, *[x.repr_value(q) for x in lv]], fill=sub_fill, font=label_font,
                fmt="0.0000", align=center)
            rc = ["", ""] + [f"{e * 100:.2f}%" for e in self.rel_change.get(q, [])]
            row(["  rel. change vs. fewer", *rc[1:]], align=center)
            p, g = self.observed_order.get(q), self.gci_fine.get(q)
            extra = []
            if p is not None:
                extra.append(f"observed order p = {p:.2f}")
            if g is not None:
                extra.append(f"GCI(fine) = {g * 100:.2f}%")
            if extra:
                row(["  " + "; ".join(extra)])
        ws.append([])

        rec = self.recommendation()
        sp = rec["speedup_vs_finest"]
        sp_txt = f" (~{sp:.1f}x faster than the finest)" if sp and sp > 1 else ""
        r = row([f"RECOMMENDATION: {rec['n_levels']} vertical levels "
                 f"(dz~{rec['dz']:.3f} m)"
                 + (f", runtime ~{rec['level'].runtime_s:.0f} s{sp_txt}"
                    if rec['level'].runtime_s else "")],
                fill=rec_fill if rec["converged"] else warn_fill, font=label_font)
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=ncol)
        r = row([rec["reason"]])
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=ncol)

        ws.column_dimensions["A"].width = 28
        for c in range(2, ncol + 1):
            ws.column_dimensions[chr(64 + c)].width = 13
        wb.save(path)
        return path


# --------------------------------------------------------------------------- #
# orchestration
# --------------------------------------------------------------------------- #

def layer_levels(cfg: Config, data: dict, counts=None) -> list[int]:
    """Layer counts to test, ascending (few -> many).

    With explicit *counts* they are used as-is; otherwise a ladder is derived from
    the inferred baseline so the layer thickness roughly halves each step (a
    refinement ratio ~2 for the GCI): ~half, baseline, ~double and ~triple the
    baseline number of intervals."""
    if counts:
        return sorted({int(c) for c in counts})
    base = threed.infer_vertical_layers(cfg, data).n_levels
    m = max(2, base - 1)                                  # intervals at the baseline
    intervals = sorted({max(2, m // 2), m, 2 * m, 3 * m})
    return [i + 1 for i in intervals]


def _build_report(levels, quantities, tolerance, probes, case="") \
        -> VerticalConvergenceReport:
    rep = VerticalConvergenceReport(
        levels=list(levels), quantities=list(quantities), tolerance=tolerance,
        probes=np.asarray(probes), case=case)
    for q in quantities:
        rep.rel_change[q] = [_rel_change(levels[i].qoi[q], levels[i + 1].qoi[q])
                             for i in range(len(levels) - 1)]
        g = _gci_triplet(levels, q)
        rep.observed_order[q], rep.gci_fine[q] = g.p, g.gci_fine
    rep.converged = all(rep.rel_change[q] and rep.rel_change[q][-1] < tolerance
                        for q in quantities)
    return rep


def make_telemac3d_simulator(cfg: Config, *, results_2d: Path, base_dir: Path,
                             runtime=None, n_processors: int | None = None,
                             total_time_factor: float = 4.0,
                             hotstart: str = "constant_depth",
                             n_time_steps: int | None = None,
                             target_time: float | None = None,
                             min_depth: float | None = None):
    """Build a ``simulate(n_levels, probes, quantities) -> VerticalLevel``.

    Each layer count runs in its own ``base_dir/L<n>/`` folder (the fixed geometry
    and boundary files are symlinked in, so the horizontal mesh is *not* rebuilt),
    launches ``telemac3d.py`` once and samples the depth-averaged QoI from the run's
    2D result.

    *target_time* (s), when set, holds the **simulated physical time constant** across
    all layer counts: each level's ``NUMBER OF TIME STEPS`` is scaled to
    ``round(target_time / dt_level)`` (``dt`` shrinks with the layer count for a fixed
    Courant, so a finer grid takes proportionally more steps). This removes the
    fixed-step-count confound where finer levels reach a much *earlier* transient state
    (and so appear "not converged" for reasons of unequal spin-up, not resolution). It
    overrides *n_time_steps*."""
    import copy
    import time

    from hydromate.env import TelemacRuntime
    from hydromate.workflow import run_solver_streaming

    rt = runtime or TelemacRuntime(cfg.telemac)
    model = Path(cfg.model_dir)
    base_dir = Path(base_dir)
    # the built 2D steering carries the PRESCRIBED FLOWRATES/ELEVATIONS/VELOCITY
    # PROFILES that build_3d_cas copies into each 3D run. It lives in the ORIGINAL
    # model dir; since each level runs in its own L<n>/ subdir (lc.model_dir = d),
    # build_3d_cas's default (lc.model_dir / cas_file) would miss it and silently drop
    # the prescribed inflow discharge - leaving the inflow unforced. Pin it explicitly.
    cas_2d_src = model / cfg.cas_file

    def simulate(n_levels, probes, quantities):
        d = base_dir / f"L{n_levels}"
        d.mkdir(parents=True, exist_ok=True)
        # the 3D case reads the same horizontal mesh + boundary file as the 2D case;
        # a user_fortran/ dir (e.g. the KB15 Spalart-Allmaras source-term fix) is
        # symlinked in too so FORTRAN FILE resolves and the patch compiles per run
        links = [cfg.geometry_slf, cfg.boundary_cli]
        if (model / "user_fortran").is_dir():
            links.append("user_fortran")
        # the continuation hotstart reads FILE FOR 2D CONTINUATION (r2d.slf) from
        # the run dir, so it must be linked in too
        if hotstart == "continuation" and Path(results_2d).name == cfg.results_slf:
            links.append(cfg.results_slf)
        for fname in links:
            src, dst = model / fname, d / fname
            if src.exists() and not dst.exists():
                dst.symlink_to(src)
        lc = copy.deepcopy(cfg)
        lc.model_dir = d
        # hotstart: "constant_depth" (default) is a robust uniform seed so every
        # per-layer run reaches steady state; "continuation" lifts the equilibrium 2D
        # field for a far shorter spin-up (paired with an explicit n_time_steps cap).
        # The seed only affects the transient, not the converged depth-averaged QoI.
        steps = n_time_steps
        if target_time is not None:
            # dt is sized from the Courant target + cell sizes, independent of the step
            # count, so build once to read it, then scale steps to hold time constant.
            probe = threed.build_3d_cas(lc, results_2d=results_2d, cas_2d=cas_2d_src,
                                        n_levels=n_levels,
                                        total_time_factor=total_time_factor,
                                        hotstart=hotstart, n_time_steps=n_time_steps,
                                        min_depth=min_depth)
            steps = max(1, round(float(target_time) / probe.time_step))
            log.info("  L%d: dt~%.4f s -> %d steps to reach %.0f s simulated time",
                     n_levels, probe.time_step, steps, float(target_time))
        setup = threed.build_3d_cas(lc, results_2d=results_2d, cas_2d=cas_2d_src,
                                    n_levels=n_levels,
                                    total_time_factor=total_time_factor,
                                    hotstart=hotstart, n_time_steps=steps,
                                    min_depth=min_depth)
        t0 = time.time()
        # stream the telemac3d listing live with the simulated-time progress bar
        # (same look as initial_run.py); 3D has no DURATION keyword, so the bar's
        # end time is the fixed step count times the per-layer time step
        proc = run_solver_streaming(
            rt, lc, cas_file=setup.cas.name, solver="telemac3d", cwd=d,
            ncsize=(n_processors or cfg.telemac.n_processors),
            duration=setup.time_step * setup.n_time_steps)
        runtime_s = time.time() - t0
        if proc.returncode != 0:
            raise RuntimeError(
                f"telemac3d failed for {n_levels} levels (rc={proc.returncode}); "
                f"see the listing under {d}")
        res2d = d / lc.results2d_from_3d_slf
        data = selafin.read_slf(res2d, frame=-1)
        return VerticalLevel(
            n_levels=n_levels, dz=setup.dz,
            n_nodes3d=data["x"].size * n_levels,
            n_elem3d=data["ikle"].shape[0] * (n_levels - 1),
            time_step=setup.time_step, runtime_s=runtime_s,
            qoi=extract_at_points(res2d, quantities, probes))

    return simulate


def run_vertical_convergence(cfg: Config, *, results_2d: str | Path | None = None,
                             counts=None, quantities=DEFAULT_QUANTITIES, probes=None,
                             tolerance: float = 0.02, simulate=None,
                             base_dir: Path | None = None, runtime=None,
                             n_processors: int | None = None,
                             total_time_factor: float = 4.0,
                             hotstart: str = "constant_depth",
                             n_time_steps: int | None = None,
                             target_time: float | None = None,
                             min_depth: float | None = None
                             ) -> VerticalConvergenceReport:
    """Run the vertical-layer convergence study and return the report.

    Needs the converged 2D result (``initial_run`` output) for the hotstart, the
    horizontal mesh and the probe points. *counts* overrides the auto-derived layer
    ladder; *simulate* overrides the TELEMAC runner (used in tests).

    *target_time* (s) holds the simulated physical time constant across layer counts
    (steps scaled per level's ``dt``), removing the fixed-step-count spin-up confound;
    it overrides *n_time_steps*."""
    model = Path(cfg.model_dir)
    results_2d = Path(results_2d) if results_2d else model / cfg.results_slf
    missing = (
        f"2D result not found: {results_2d}. Run preprocessing.py + initial_run.py "
        "first (the vertical study hotstarts the 3D runs from the 2D result).")
    if counts:                                       # explicit ladder: no 2D read needed
        counts = sorted({int(c) for c in counts})
    else:
        if not results_2d.exists():
            raise FileNotFoundError(missing)
        counts = layer_levels(cfg, selafin.read_slf(results_2d, frame=-1))
    quantities = list(quantities)
    probes = np.asarray(probes if probes is not None else default_probes(cfg), float)
    if simulate is None:                             # real runs need the 2D result + dirs
        if not results_2d.exists():
            raise FileNotFoundError(missing)
        if base_dir is None:
            base_dir = cfg.postprocessing_path("vertical-convergence")
        simulate = make_telemac3d_simulator(
            cfg, results_2d=results_2d, base_dir=base_dir, runtime=runtime,
            n_processors=n_processors, total_time_factor=total_time_factor,
            hotstart=hotstart, n_time_steps=n_time_steps, target_time=target_time,
            min_depth=min_depth)

    results: list[VerticalLevel] = []
    for n in counts:
        log.info("vertical convergence: running %d sigma levels", n)
        results.append(simulate(n, probes, quantities))
        r = results[-1]
        log.info("  %d levels: dz~%.3f m, %d 3D nodes, dt~%s s, runtime %s",
                 n, r.dz, r.n_nodes3d, r.time_step,
                 f"{r.runtime_s:.0f}s" if r.runtime_s else "?")
    rep = _build_report(results, quantities, tolerance, probes,
                        case=getattr(cfg, "name", ""))
    rec = rep.recommendation()
    log.info("vertical convergence: %s at %.1f%% tolerance; recommended %d levels",
             "CONVERGED" if rep.converged else "NOT converged", tolerance * 100,
             rec["n_levels"])
    return rep
