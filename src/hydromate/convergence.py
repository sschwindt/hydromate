"""Mesh-convergence (grid-independence) study for the 2D hydraulic case.

Runs the *same* steady simulation on a sequence of progressively finer meshes and
monitors quantities of interest phi (water depth h, scalar velocity U) at a fixed
set of probe points. Convergence is judged on the Grid Convergence Index (GCI)
of the finest grid triplet - with an asymptotic-range check to decide whether
the GCI is trustworthy, else on the finest successive relative change - against
a tolerance (default 5%; Celik et al. 2008 / Roache: GCI ~5-10% is acceptable
engineering accuracy, ~1-2% is high-precision work). This is the formal way to
quantify the spatial-discretization error and demonstrate grid independence.

Notes following standard practice (Celik et al. 2008, ASME J. Fluids Eng. 130):

* Mesh convergence is distinct from iterative (solver) convergence - the latter
  (residual reduction within one run) must be reached first or the comparison is
  meaningless. The runs here use TELEMAC's variable time step so each mesh marches
  at its own CFL-admissible dt (refining Delta x forces a smaller dt).
* Use at least three meshes with a refinement factor **r >= 1.3** between grids
  (:func:`ratio_levels`). Closer-spaced grids push the grid-to-grid solution
  differences into iterative/sampling noise and degenerate the GCI denominator
  ``r**p - 1``, making the observed order p and the GCI unstable.
* Every level (and every candidate extension) gets the physical validity checks
  of :mod:`hydromate.mesh_validity` (y+, roughness regime, dx vs. ks, turbulence
  model consistency) - refinement is only meaningful while the closures hold.

The study can **resume** from completed per-level results (each level folder
carries a ``level.json``) and can **extend** the ladder one finer mesh at a time
after showing a runtime estimate + validity check (interactive prompt; in a
non-interactive session it stops and reports instead).

The TELEMAC execution is injected (the ``simulate`` callable), so the report
maths are unit-testable without the solver.
"""

from __future__ import annotations

import json
import logging
import math
from dataclasses import dataclass, field
from pathlib import Path

import numpy as np

from hydromate import selafin
from hydromate.config import Config

log = logging.getLogger("hydromate")

DEFAULT_QUANTITIES = ("WATER DEPTH", "SCALAR VELOCITY")
GCI_SAFETY = 1.25        # Roache factor of safety for >=3 grids
DEFAULT_TOLERANCE = 0.05  # GCI / relative-change goal (Celik et al. 2008: 5-10%)
THEORETICAL_ORDER = 2.0  # TELEMAC's FE scheme is formally <= 2nd order; the
#                          observed p is clamped to this for the GCI so a
#                          noise-driven p (e.g. 7) cannot fake a tiny GCI
ASYMPTOTIC_BAND = (0.8, 1.25)  # GCI ratio band within which the triplet is taken
#                                to be in the asymptotic range (GCI trustworthy)


# --------------------------------------------------------------------------- #
# quantity-of-interest extraction from a result SELAFIN
# --------------------------------------------------------------------------- #


def _resolve_field(values: dict[str, np.ndarray], qty: str) -> np.ndarray:
    """Node field for *qty*, deriving it from components/aliases when needed."""
    names = {k.upper(): k for k in values}
    q = qty.upper()
    if q in names:
        return values[names[q]]
    if q == "SCALAR VELOCITY":
        u = next((values[names[n]] for n in ("VELOCITY U", "U") if n in names), None)
        v = next((values[names[n]] for n in ("VELOCITY V", "V") if n in names), None)
        if u is not None and v is not None:
            return np.hypot(u, v)
    if q == "WATER DEPTH" and "FREE SURFACE" in names and "BOTTOM" in names:
        return values[names["FREE SURFACE"]] - values[names["BOTTOM"]]
    raise KeyError(f"quantity {qty!r} not in result (have {list(values)})")


def _interp_to_points(x, y, field, points):
    """Linear interpolation of a node field onto *points*; nearest for outliers."""
    from scipy.interpolate import LinearNDInterpolator, NearestNDInterpolator

    nodes = np.column_stack([x, y])
    out = LinearNDInterpolator(nodes, field)(points)
    nan = np.isnan(out)
    if nan.any():
        out[nan] = NearestNDInterpolator(nodes, field)(points[nan])
    return out


def extract_at_points(slf_path, quantities, points) -> dict[str, np.ndarray]:
    """Read the last frame of *slf_path* and sample *quantities* at *points*."""
    data = selafin.read_slf(slf_path, frame=-1)
    points = np.asarray(points, dtype=float)
    return {q: _interp_to_points(data["x"], data["y"], _resolve_field(data["values"], q),
                                 points) for q in quantities}


# --------------------------------------------------------------------------- #
# data model
# --------------------------------------------------------------------------- #


@dataclass
class LevelResult:
    """One mesh in the sequence: its resolution, size and sampled QoI."""

    label: str
    cell_size: float                       # representative (channel) edge length [m]
    n_elem: int
    n_nodes: int
    min_edge: float
    dt: float | None                       # CFL-admissible time step estimate [s]
    runtime_s: float | None
    qoi: dict[str, np.ndarray]             # {quantity: (n_probes,) values}
    scale: float = 1.0                     # mesh.size_scale this level was built at
    mean_depth: float | None = None        # mean probe water depth [m]
    mean_velocity: float | None = None     # mean probe scalar velocity [m/s]
    validity: object | None = None         # mesh_validity.MeshValidity (or None)

    def repr_value(self, qty: str) -> float:
        """Representative scalar: mean magnitude over the probes."""
        return float(np.mean(np.abs(self.qoi[qty])))


@dataclass
class ConvergenceReport:
    levels: list[LevelResult]              # ordered coarse -> fine
    quantities: list[str]
    tolerance: float
    probes: np.ndarray
    rel_change: dict[str, list[float]] = field(default_factory=dict)   # per qty, len n-1
    observed_order: dict[str, float | None] = field(default_factory=dict)
    gci_fine: dict[str, float | None] = field(default_factory=dict)
    converged: bool = False
    case: str = ""                          # case name (for the report title)
    order_used: dict[str, float | None] = field(default_factory=dict)  # p clamped for GCI
    gci_med: dict[str, float | None] = field(default_factory=dict)     # medium-grid GCI
    asymptotic_ratio: dict[str, float | None] = field(default_factory=dict)
    verdict_basis: str = "rel_change"       # "gci" | "rel_change"

    def format(self) -> str:
        lv = self.levels
        w = max(12, *(len(x.label) for x in lv))
        cols = "".join(f"{x.label:>{w}}" for x in lv)
        lines = [
            "Mesh-convergence study (coarse -> fine)",
            f"  probes: {len(self.probes)}   tolerance: {self.tolerance * 100:.1f}%",
            f"  {'cell size [m]':<22}" + "".join(f"{x.cell_size:>{w}.3f}" for x in lv),
            f"  {'size scale [-]':<22}" + "".join(f"{x.scale:>{w}.3f}" for x in lv),
            f"  {'elements':<22}" + "".join(f"{x.n_elem:>{w}d}" for x in lv),
            f"  {'nodes':<22}" + "".join(f"{x.n_nodes:>{w}d}" for x in lv),
            f"  {'shortest edge [m]':<22}" + "".join(f"{x.min_edge:>{w}.4f}" for x in lv),
            f"  {'time step dt [s]':<22}"
            + "".join((f"{x.dt:>{w}.4f}" if x.dt else f"{'-':>{w}}") for x in lv),
            f"  {'runtime [s]':<22}"
            + "".join((f"{x.runtime_s:>{w}.1f}" if x.runtime_s else f"{'-':>{w}}")
                      for x in lv),
            "  " + "-" * (22 + w * len(lv)),
        ]
        lines.append(f"  {'quantity phi':<22}" + cols)
        for q in self.quantities:
            row = "".join(f"{x.repr_value(q):>{w}.4f}" for x in lv)
            lines.append(f"  {q:<22}{row}")
            # relative change between successive meshes, aligned under the finer one
            rc = self.rel_change.get(q, [])
            cells = [f"{'':>{w}}"] + [f"{e * 100:>{w - 1}.2f}%" for e in rc]
            lines.append(f"    {'rel. change':<20}" + "".join(cells))
            extra = []
            p, pu = self.observed_order.get(q), self.order_used.get(q)
            if p is not None:
                extra.append(f"observed order p={p:.2f}"
                             + (f" (p_used={pu:.2f})"
                                if pu is not None and abs(pu - p) > 1e-9 else ""))
            if self.gci_fine.get(q) is not None:
                extra.append(f"GCI(fine)={self.gci_fine[q] * 100:.2f}%")
            if self.gci_med.get(q) is not None:
                extra.append(f"GCI(med)={self.gci_med[q] * 100:.2f}%")
            if self.asymptotic_ratio.get(q) is not None:
                extra.append(f"asymptotic ratio={self.asymptotic_ratio[q]:.2f}")
            if extra:
                lines.append(f"    {'':<20}" + "; ".join(extra))
        verdict = "CONVERGED" if self.converged else "NOT converged"
        basis = ("GCI of the finest grid triplet" if self.verdict_basis == "gci"
                 else "finest relative change; GCI not trustworthy - asymptotic "
                      f"range check outside {ASYMPTOTIC_BAND[0]}..{ASYMPTOTIC_BAND[1]} "
                      "or order not computable")
        lines.append(f"  => {verdict} at {self.tolerance * 100:.1f}% ({basis})")
        if not self.converged:
            lines.append("     refine further (the finest mesh still changes the QoI "
                         "by more than the tolerance).")
        validity = [x for x in lv if x.validity is not None]
        if validity:
            lines.append("  mesh validity (y+ / roughness / turbulence closure):")
            for x in validity:
                vlines = x.validity.lines()
                lines.append(f"    {x.label}: {vlines[0]}")
                lines.extend(f"      {extra_line}" for extra_line in vlines[1:])
        return "\n".join(lines)

    def save(self, path) -> Path:
        path = Path(path)
        path.write_text(self.format() + "\n")
        return path

    def recommendation(self) -> dict:
        """Recommend the coarsest grid-independent cell size (convergence vs. cost).

        A mesh is adequate when refining it further (to the next finer mesh)
        changes every quantity of interest by less than the tolerance; the
        coarsest such mesh is the cheapest converged choice. Returns a dict with
        the recommended :class:`LevelResult`, its cell size, whether it converged,
        a one-line reason, and the runtime speed-up versus the finest mesh.
        """
        lv = self.levels
        chosen = None
        for i in range(len(lv) - 1):                       # coarse -> fine
            if all(self.rel_change[q][i] < self.tolerance for q in self.quantities):
                chosen = lv[i]
                break
        if chosen is not None:
            converged = True
            reason = (f"coarsest mesh whose water depth and velocity stay within "
                      f"{self.tolerance * 100:.0f}% under refinement (grid-independent "
                      "and cheapest to run)")
        elif self.verdict_basis == "gci" and self.converged and len(lv) >= 2:
            # GCI-converged although no single level passed the per-level scan:
            # the medium grid of the converged fine triplet is the cheapest mesh
            # whose discretization error is bounded (by GCI(med))
            chosen = lv[-2]
            converged = True
            gm = max((self.gci_med.get(q) or 0.0) for q in self.quantities)
            reason = (f"medium grid of the GCI-converged fine triplet "
                      f"(discretization error bounded by GCI(med) = {gm * 100:.2f}%)")
        else:
            chosen = lv[-1]
            converged = False
            reason = (f"no mesh met the {self.tolerance * 100:.0f}% tolerance; using "
                      "the finest (refine further to confirm grid independence)")
        finest = lv[-1]
        speedup = (finest.runtime_s / chosen.runtime_s
                   if chosen.runtime_s and finest.runtime_s else None)
        return {"level": chosen, "cell_size": chosen.cell_size,
                "converged": converged, "reason": reason, "speedup_vs_finest": speedup}

    def to_xlsx(self, path) -> Path:
        """Write a styled .xlsx convergence report with the cell-size recommendation."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        path = Path(path)
        lv = self.levels
        head_fill = PatternFill("solid", fgColor="1F4E78")     # dark blue
        head_font = Font(bold=True, color="FFFFFF")
        label_font = Font(bold=True)
        rec_fill = PatternFill("solid", fgColor="C6EFCE")      # green
        warn_fill = PatternFill("solid", fgColor="FFC7CE")     # red
        sub_fill = PatternFill("solid", fgColor="DDEBF7")      # light blue
        center = Alignment(horizontal="center")
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        wb = Workbook()
        ws = wb.active
        ws.title = "mesh convergence"
        ncol = 1 + len(lv)

        def row(cells, *, fill=None, font=None, fmt=None, align=None):
            ws.append(list(cells))
            r = ws.max_row
            for c in range(1, len(cells) + 1):
                cell = ws.cell(row=r, column=c)
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font
                if fmt and c > 1 and isinstance(cell.value, (int, float)):
                    cell.number_format = fmt
                if align and c > 1:
                    cell.alignment = align
                cell.border = border
            return r

        # title
        title = "Mesh-convergence study"
        if getattr(self, "case", ""):
            title += f" - case '{self.case}'"
        ws.append([title])
        ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=ncol)
        ws["A1"].font = Font(bold=True, size=14)
        basis = ("GCI-based" if self.verdict_basis == "gci"
                 else "relative-change fallback (GCI not trustworthy)")
        ws.append([f"probes: {len(self.probes)}    tolerance: {self.tolerance * 100:.1f}%    "
                   f"meshes: {len(lv)} (coarse to fine)    verdict: {basis}"])
        ws.merge_cells(start_row=2, end_row=2, start_column=1, end_column=ncol)
        ws.append([])

        row(["", *[x.label for x in lv]], fill=head_fill, font=head_font, align=center)
        row(["cell size [m]", *[x.cell_size for x in lv]], font=label_font,
            fmt="0.000", align=center)
        row(["size scale [-]", *[x.scale for x in lv]], font=label_font,
            fmt="0.000", align=center)
        row(["elements", *[x.n_elem for x in lv]], font=label_font, fmt="#,##0",
            align=center)
        row(["nodes", *[x.n_nodes for x in lv]], font=label_font, fmt="#,##0",
            align=center)
        row(["shortest edge [m]", *[x.min_edge for x in lv]], font=label_font,
            fmt="0.0000", align=center)
        row(["time step dt [s]", *[(x.dt or float('nan')) for x in lv]],
            font=label_font, fmt="0.0000", align=center)
        row(["runtime [s]", *[(x.runtime_s or float('nan')) for x in lv]],
            font=label_font, fmt="0.0", align=center)
        ws.append([])

        # physical validity of the resolution (see hydromate.mesh_validity)
        vals = [x.validity for x in lv]
        if any(v is not None for v in vals):
            def vattr(attr, absent="-"):
                out = []
                for v in vals:
                    a = getattr(v, attr, None) if v is not None else None
                    out.append(a if (a is not None
                                     and not (isinstance(a, float) and math.isnan(a)))
                               else absent)
                return out

            row(["mean depth at probes [m]", *vattr("depth")], font=label_font,
                fmt="0.00", align=center)
            row(["mean velocity [m/s]", *vattr("velocity")], font=label_font,
                fmt="0.00", align=center)
            row(["u* (log law) [m/s]", *vattr("u_star")], font=label_font,
                fmt="0.000", align=center)
            r = row(["wall y+", *vattr("y_plus")], font=label_font, fmt="0",
                    align=center)
            for c, v in enumerate(vals, start=2):
                if v is not None and not v.y_plus_ok:
                    ws.cell(row=r, column=c).fill = warn_fill
            r = row(["ks+ (roughness Re)", *vattr("ks_plus")], font=label_font,
                    fmt="0", align=center)
            for c, v in enumerate(vals, start=2):
                if v is not None and v.ks_regime not in ("fully rough", "unknown"):
                    ws.cell(row=r, column=c).fill = warn_fill
            r = row(["cell size / ks [-]", *vattr("dx_over_ks")], font=label_font,
                    fmt="0.00", align=center)
            for c, v in enumerate(vals, start=2):
                if v is not None and v.dx_below_ks:
                    ws.cell(row=r, column=c).fill = warn_fill
            r = row(["turbulence pick (auto)",
                     *[(v.turbulence_pick_name or "-") if v is not None else "-"
                       for v in vals]], font=label_font, align=center)
            for c, v in enumerate(vals, start=2):
                if v is not None and v.regime_change:
                    ws.cell(row=r, column=c).fill = warn_fill
            warned = [w for v in vals if v is not None for w in v.warnings]
            for w in dict.fromkeys(warned):              # unique, order-preserving
                row(["  ! " + w])
            ws.append([])

        for q in self.quantities:
            row([q, *[x.repr_value(q) for x in lv]], fill=sub_fill, font=label_font,
                fmt="0.0000", align=center)
            rc = ["", ""] + [f"{e * 100:.2f}%" for e in self.rel_change.get(q, [])]
            row(["  rel. change vs. coarser", *rc[1:]], align=center)
            p, pu = self.observed_order.get(q), self.order_used.get(q)
            extra = []
            if p is not None:
                extra.append(f"observed order p = {p:.2f}"
                             + (f" (p_used = {pu:.2f})"
                                if pu is not None and abs(pu - p) > 1e-9 else ""))
            if self.gci_fine.get(q) is not None:
                extra.append(f"GCI(fine) = {self.gci_fine[q] * 100:.2f}%")
            if self.gci_med.get(q) is not None:
                extra.append(f"GCI(med) = {self.gci_med[q] * 100:.2f}%")
            if self.asymptotic_ratio.get(q) is not None:
                extra.append(f"asymptotic ratio = {self.asymptotic_ratio[q]:.2f}")
            if extra:
                row(["  " + "; ".join(extra)])
        ws.append([])

        rec = self.recommendation()
        sp = rec["speedup_vs_finest"]
        sp_txt = f" (~{sp:.1f}x faster than the finest)" if sp and sp > 1 else ""
        r = row([f"RECOMMENDATION: cell size = {rec['cell_size']:.3f} m"
                 + (f", runtime ~{rec['level'].runtime_s:.0f} s{sp_txt}"
                    if rec['level'].runtime_s else "")],
                fill=rec_fill if rec["converged"] else warn_fill, font=label_font)
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=ncol)
        r = row([rec["reason"]])
        ws.merge_cells(start_row=r, end_row=r, start_column=1, end_column=ncol)

        ws.column_dimensions["A"].width = 28
        for c in range(2, ncol + 1):
            ws.column_dimensions[chr(64 + c)].width = 13
        wb.save(path)
        return path


# --------------------------------------------------------------------------- #
# report maths
# --------------------------------------------------------------------------- #


def _rel_change(prev: np.ndarray, curr: np.ndarray) -> float:
    """Relative L2 change between two probe vectors, normalised by the finer one."""
    denom = np.linalg.norm(curr)
    if denom == 0:
        return float(np.linalg.norm(curr - prev))
    return float(np.linalg.norm(curr - prev) / denom)


def build_report(levels: list[LevelResult], quantities, tolerance: float,
                 probes: np.ndarray) -> ConvergenceReport:
    """Assemble a :class:`ConvergenceReport` (relative change, order, GCI, verdict).

    The verdict is **GCI-based** when, for every quantity, the finest grid triplet
    yields a GCI whose asymptotic-range check passes (``GCI(med) / (r**p GCI(fine))``
    within :data:`ASYMPTOTIC_BAND`, ~1 in the asymptotic range): converged iff
    ``GCI(fine) < tolerance``. Otherwise it falls back to the finest successive
    relative change vs. the tolerance (recorded in ``verdict_basis``).
    """
    rep = ConvergenceReport(levels=list(levels), quantities=list(quantities),
                            tolerance=tolerance, probes=np.asarray(probes))
    for q in quantities:
        changes = [_rel_change(levels[i].qoi[q], levels[i + 1].qoi[q])
                   for i in range(len(levels) - 1)]
        rep.rel_change[q] = changes
        g = _gci_triplet(levels, q)
        rep.observed_order[q] = g.p
        rep.order_used[q] = g.p_used
        rep.gci_fine[q] = g.gci_fine
        rep.gci_med[q] = g.gci_med
        rep.asymptotic_ratio[q] = g.asymptotic_ratio

    def trustworthy(q):
        ar = rep.asymptotic_ratio.get(q)
        return (rep.gci_fine.get(q) is not None and ar is not None
                and ASYMPTOTIC_BAND[0] <= ar <= ASYMPTOTIC_BAND[1])

    if quantities and all(trustworthy(q) for q in quantities):
        rep.verdict_basis = "gci"
        rep.converged = all(rep.gci_fine[q] < tolerance for q in quantities)
    else:
        rep.verdict_basis = "rel_change"
        rep.converged = all(rep.rel_change[q] and rep.rel_change[q][-1] < tolerance
                            for q in quantities)
    return rep


@dataclass
class _Gci:
    """Order/GCI diagnostics of one quantity's finest grid triplet."""

    p: float | None = None                 # observed order (raw)
    p_used: float | None = None            # p clamped to [0.5, THEORETICAL_ORDER]
    gci_fine: float | None = None          # fine-grid GCI (medium->fine pair)
    gci_med: float | None = None           # medium-grid GCI (coarse->medium pair)
    asymptotic_ratio: float | None = None  # GCI(med) / (r2**p_used * GCI(fine))


def _gci_triplet(levels: list[LevelResult], q: str) -> _Gci:
    """Observed order and GCIs from the **finest** (>=3) grid triplet.

    The GCI uses the observed order clamped to ``[0.5, THEORETICAL_ORDER]``: with
    closely spaced or noisy grids the raw p can come out absurdly high (the ering
    case produced p=7 from re-run noise), and ``r**p - 1`` in the denominator then
    shrinks the GCI to a meaninglessly small value. The raw p is still reported.
    """
    g = _Gci()
    if len(levels) < 3:
        return g
    coarse, med, fine = levels[-3], levels[-2], levels[-1]
    r1 = coarse.cell_size / med.cell_size
    r2 = med.cell_size / fine.cell_size
    if min(r1, r2) <= 1.0:
        return g
    f_c, f_m, f_f = coarse.repr_value(q), med.repr_value(q), fine.repr_value(q)
    e21, e32 = f_m - f_c, f_f - f_m
    if e21 == 0 or e32 == 0 or (e32 / e21) <= 0:
        return g                              # oscillatory / degenerate: no order
    p = math.log(abs(e21 / e32)) / math.log(0.5 * (r1 + r2))
    if not math.isfinite(p):
        return g
    g.p = p
    if p <= 0:
        return g
    g.p_used = min(max(p, 0.5), THEORETICAL_ORDER)
    rel_f = abs(e32 / f_f) if f_f != 0 else abs(e32)
    rel_m = abs(e21 / f_m) if f_m != 0 else abs(e21)
    g.gci_fine = GCI_SAFETY * rel_f / (r2 ** g.p_used - 1.0)
    g.gci_med = GCI_SAFETY * rel_m / (r1 ** g.p_used - 1.0)
    if g.gci_fine > 0:
        g.asymptotic_ratio = g.gci_med / (r2 ** g.p_used * g.gci_fine)
    return g


# --------------------------------------------------------------------------- #
# orchestration
# --------------------------------------------------------------------------- #


def default_levels(cfg: Config) -> list[tuple[str, float, float]]:
    """Three meshes spanning one doubling of cell size around the config sizes.

    Returns ``(label, channel_size, floodplain_size)`` coarse -> fine. The finest
    matches the config; coarser ones double the cell size (so coarse/fine = 2)."""
    ch, fp = cfg.mesh.channel_size, cfg.mesh.floodplain_size
    return [("coarse", ch * 2.0, fp * 2.0),
            ("medium", ch * math.sqrt(2.0), fp * math.sqrt(2.0)),
            ("fine", ch, fp)]


def percent_levels(cfg: Config,
                   percents=(0.40, 0.20, 0.0, -0.20, -0.40)
                   ) -> list[tuple[str, float, float]]:
    """Cell-size levels at fractional offsets from the config size, coarse -> fine.

    Positive *percents* coarsen (larger cells), negative refine (smaller cells);
    the default gives two 40%/20%-coarser, the baseline, and two 20%/40%-finer
    meshes. Returns ``(label, channel_size, floodplain_size)``.

    .. note:: prefer :func:`ratio_levels` - the +-20% steps give successive grid
       ratios of only ~1.17-1.33, below the r >= 1.3 that Celik et al. (2008)
       recommend for a stable observed order / GCI."""
    ch, fp = cfg.mesh.channel_size, cfg.mesh.floodplain_size
    levels = []
    for p in sorted(percents, reverse=True):           # coarse (large) -> fine (small)
        factor = 1.0 + p
        if abs(p) < 1e-9:
            label = "baseline"
        elif p > 0:
            label = f"+{p * 100:.0f}% coarser"
        else:
            label = f"-{-p * 100:.0f}% finer"
        levels.append((label, ch * factor, fp * factor))
    return levels


def ratio_levels(cfg: Config, ratio: float = 1.3, n_coarser: int = 1,
                 n_finer: int = 2) -> list[tuple[str, float, float]]:
    """Cell-size ladder with a constant refinement factor *ratio*, coarse -> fine.

    The default gives four meshes: one coarser (``x ratio``), the config baseline,
    and two finer (``/ ratio``, ``/ ratio**2``) - the finest three form the GCI
    triplet. Celik et al. (2008) recommend **ratio >= 1.3**: with closer-spaced
    grids the solution differences drop into iterative/sampling noise and the GCI
    denominator ``r**p - 1`` degenerates, so p and the GCI become unstable.
    Returns ``(label, channel_size, floodplain_size)`` tuples; extend the ladder
    one finer level at a time via ``run_mesh_convergence(extend_ratio=...)``.
    """
    if ratio <= 1.0:
        raise ValueError(f"refinement ratio must exceed 1.0 (got {ratio}); "
                         "Celik et al. (2008) recommend >= 1.3")
    ch, fp = cfg.mesh.channel_size, cfg.mesh.floodplain_size
    levels = []
    for k in range(n_coarser, -(n_finer + 1), -1):     # coarse (large) -> fine (small)
        f = ratio ** k
        if k == 0:
            label = "baseline"
        elif k > 0:
            label = f"coarser x{f:.2f}"
        else:
            label = f"finer x{1.0 / f:.2f}"
        levels.append((label, ch * f, fp * f))
    return levels


def _next_finer(levels: list[tuple[str, float, float]],
                ratio: float) -> tuple[str, float, float]:
    """The next extension level: the finest tuple refined by *ratio*, labelled by
    its total refinement factor vs. the baseline (or coarsest) level."""
    _, ch0, fp0 = levels[-1]
    ch, fp = ch0 / ratio, fp0 / ratio
    base = next((c for lbl, c, _ in levels if lbl == "baseline"), levels[0][1])
    return f"finer x{base / ch:.2f}", ch, fp


def _slug(label: str) -> str:
    """Filesystem-safe per-level folder name (labels carry spaces/%/x/+/- that
    TELEMAC and shells dislike in paths)."""
    return "".join(c if c.isalnum() else "_" for c in label).strip("_") or "level"


def default_probes(cfg: Config, n: int = 40) -> np.ndarray:
    """Probe points for the QoI: the ground-truth points if available, else samples
    along the channel centerline (falling back to the ROI bbox diagonal)."""
    gt = cfg.ground_truth_path
    if gt is not None and Path(gt).exists():
        try:
            from hydromate import ground_truth

            tables = ground_truth.read_tidy(gt)
            pts = np.vstack([t[["x", "y"]].to_numpy(dtype=float)
                             for t in tables.values() if {"x", "y"} <= set(t.columns)])
            if len(pts):
                return pts
        except Exception as exc:  # pragma: no cover - fallback only
            log.debug("probe points from ground truth unavailable: %s", exc)
    if cfg.geodata.channel_centerline is not None:
        import geopandas as gpd
        from shapely.ops import linemerge, unary_union

        gdf = gpd.read_file(cfg.geodata.channel_centerline)
        if gdf.crs and gdf.crs.to_epsg() != cfg.crs_epsg:
            gdf = gdf.to_crs(epsg=cfg.crs_epsg)
        merged = unary_union(gdf.geometry.values)
        # linemerge only accepts a MultiLineString; a single merged LineString is
        # already the line (passing it to linemerge raises "Cannot linemerge")
        line = linemerge(merged) if merged.geom_type == "MultiLineString" else merged
        if line.geom_type == "MultiLineString":
            line = max(line.geoms, key=lambda g: g.length)
        d = np.linspace(0, line.length, n)
        return np.array([[line.interpolate(s).x, line.interpolate(s).y] for s in d])
    raise ValueError("no probe points: provide ground truth or a channel centerline, "
                     "or pass probes= explicitly to run_mesh_convergence")


def make_telemac_simulator(cfg: Config, discharge: float, *, base_dir: Path,
                           runtime=None, n_processors: int | None = None):
    """Build a ``simulate(label, ch, fp, probes, quantities) -> LevelResult`` that
    builds the case at the given resolution, runs TELEMAC once at constant
    *discharge* (variable time step for per-mesh CFL), and samples the QoI.

    The level's ``(ch, fp)`` sizes are interpreted **relative to the config
    baseline**: the level copy is built with ``mesh.size_scale = ch /
    cfg.mesh.channel_size`` so the *whole* sizing field scales - including any
    per-zone ``Max Edge Length (m)`` values from the mesh-zones gpkg, which would
    otherwise silently override the per-level sizes and leave every "level" with
    the same mesh (the bug behind the invalid first ering study).
    """
    import copy
    import time

    from hydromate import mesh as mesh_mod
    from hydromate import mesh_validity, pipeline, steering

    base_dir = Path(base_dir)
    # pin the turbulence closure to the baseline choice across all grid levels, so
    # the study isolates the discretization error (per-level auto-selection would
    # otherwise switch closures as the cell size changes and confound the comparison)
    pinned_turbulence = steering.select_turbulence_model(cfg)[0]
    base_channel = float(cfg.mesh.channel_size)
    try:
        ks_channel = mesh_validity.channel_ks(cfg)
    except Exception as exc:                     # pragma: no cover - geodata issues
        log.debug("channel ks unavailable: %s", exc)
        ks_channel = None

    def simulate(label, ch, fp, probes, quantities):
        lc = copy.deepcopy(cfg)
        # scale the whole sizing field (config *and* gpkg per-zone sizes) uniformly
        lc.mesh.size_scale = cfg.mesh.size_scale * (ch / base_channel)
        lc.hydrodynamics.turbulence_model = pinned_turbulence
        lc.boundaries.prescribed_flowrate = discharge
        # let TELEMAC pick a CFL-admissible dt for this mesh (iterative convergence
        # to steady state must be reached before meshes are compared). These runs are
        # pre-wetted (hotstart), so a faster 0.6 Courant target keeps the study
        # tractable; the production dry-start default is the safer 0.30.
        lc.hydrodynamics.variable_timestep = True
        lc.hydrodynamics.desired_courant = 0.6
        if n_processors:
            lc.telemac.n_processors = n_processors
        d = base_dir / _slug(label)
        lc.preprocessing_dir = d / "preprocessing"
        lc.model_dir = d / "model"
        lc.postprocessing_dir = d / "postprocessing"
        lc.calibration_dir = d / "calibration"
        lc.ensure_dirs()

        t0 = time.time()
        # reuse the parent run's one compound logfile (no per-level logfiles)
        pipeline.run(lc, validate_env=False, dry_run=True, log_to_file=False)
        runtime_s = time.time() - t0

        eff = mesh_mod.nominal_channel_size(lc)   # effective (gpkg-aware) size
        res = lc.model_path(lc.results_slf)
        data = selafin.read_slf(res, frame=-1)
        min_edge, mean_depth, mean_vel = _geom_stats(data, probes)
        dt = _cfl_dt(min_edge, mean_depth, mean_vel)
        qoi = extract_at_points(res, quantities, probes)
        validity = mesh_validity.check_level(
            cfg, dx=eff, depth=mean_depth, velocity=mean_vel, ks=ks_channel,
            pinned_model=pinned_turbulence)
        result = LevelResult(label=label, cell_size=eff, n_elem=data["ikle"].shape[0],
                             n_nodes=data["x"].size, min_edge=min_edge, dt=dt,
                             runtime_s=runtime_s, qoi=qoi, scale=lc.mesh.size_scale,
                             mean_depth=mean_depth, mean_velocity=mean_vel,
                             validity=validity)
        _write_level_json(d, result, ch, fp)
        return result

    return simulate


def _geom_stats(data: dict, probes):
    """(shortest edge, mean probe depth, mean probe velocity) from a result dict."""
    x, y, ikle = data["x"], data["y"], data["ikle"]
    p = ikle
    e = np.concatenate([
        np.hypot(x[p[:, 0]] - x[p[:, 1]], y[p[:, 0]] - y[p[:, 1]]),
        np.hypot(x[p[:, 1]] - x[p[:, 2]], y[p[:, 1]] - y[p[:, 2]]),
        np.hypot(x[p[:, 2]] - x[p[:, 0]], y[p[:, 2]] - y[p[:, 0]]),
    ])
    # a merged parallel result duplicates subdomain-interface nodes, producing
    # zero-length edges that would report min_edge=0 (and a NaN CFL dt)
    nonzero = e > 1e-9
    if not nonzero.all():
        log.debug("  _geom_stats: ignoring %d zero-length edges (duplicate "
                  "interface nodes in a merged parallel result)",
                  int((~nonzero).sum()))
        e = e[nonzero]
    probes = np.asarray(probes, dtype=float)
    try:
        depth = float(np.nanmean(_interp_to_points(
            x, y, _resolve_field(data["values"], "WATER DEPTH"), probes)))
        vel = float(np.nanmean(_interp_to_points(
            x, y, _resolve_field(data["values"], "SCALAR VELOCITY"), probes)))
    except KeyError:
        depth, vel = 1.0, 1.0
    min_edge = float(e.min()) if e.size else float("nan")
    return min_edge, max(depth, 1e-3), max(vel, 0.0)


def _cfl_dt(min_edge: float, depth: float, velocity: float, courant: float = 0.9) -> float:
    """CFL-admissible time step dt = C * dx / (|U| + sqrt(g*h))."""
    celerity = velocity + math.sqrt(9.81 * depth)
    return courant * min_edge / celerity if celerity > 0 else float("nan")


# --------------------------------------------------------------------------- #
# per-level persistence (resume support)
# --------------------------------------------------------------------------- #


def _write_level_json(level_dir: Path, level: LevelResult, ch: float,
                      fp: float) -> None:
    """Persist a completed level's metadata so a later run can resume from it.

    The QoI are deliberately *not* persisted - on resume they are re-extracted
    from the existing result SELAFIN, which keeps resume robust to changed probe
    points or quantities.
    """
    from datetime import datetime, timezone

    meta = {"label": level.label, "channel_size": ch, "floodplain_size": fp,
            "scale": level.scale, "cell_size": level.cell_size,
            "n_elem": level.n_elem, "n_nodes": level.n_nodes,
            "min_edge": level.min_edge, "dt": level.dt,
            "runtime_s": level.runtime_s,
            "created": datetime.now(timezone.utc).isoformat(timespec="seconds")}
    try:
        (Path(level_dir) / "level.json").write_text(json.dumps(meta, indent=2) + "\n")
    except OSError as exc:                       # pragma: no cover - fs issues
        log.warning("could not write %s/level.json: %s", level_dir, exc)


def _scan_completed(cfg: Config, base_dir: Path, levels) -> dict[str, Path]:
    """Map ``{label: level_dir}`` of planned levels with a reusable result.

    A level is reusable when its folder holds both the result SELAFIN and a
    ``level.json`` whose channel/floodplain sizes match the planned tuple (so a
    changed ladder never silently reuses a mesh built at different sizes)."""
    results_slf = getattr(cfg, "results_slf", "r2d.slf")
    done: dict[str, Path] = {}
    for label, ch, fp in levels:
        d = Path(base_dir) / _slug(label)
        jf = d / "level.json"
        if not (jf.exists() and (d / "model" / results_slf).exists()):
            continue
        try:
            meta = json.loads(jf.read_text())
            if (math.isclose(float(meta["channel_size"]), ch, rel_tol=1e-6)
                    and math.isclose(float(meta["floodplain_size"]), fp, rel_tol=1e-6)):
                done[label] = d
            else:
                log.info("existing level '%s' was built at different sizes "
                         "(%.3f/%.3f m vs. planned %.3f/%.3f m); it will be re-run",
                         label, meta["channel_size"], meta["floodplain_size"], ch, fp)
        except (KeyError, ValueError, TypeError, OSError) as exc:
            log.debug("unreadable level.json in %s: %s", d, exc)
    return done


def _load_completed_level(cfg: Config, level_dir: Path, label: str, probes,
                          quantities, *, ks=None,
                          pinned_model: int | None = None) -> LevelResult | None:
    """Rebuild a :class:`LevelResult` from an existing per-level folder, or None
    (-> the level is re-run). QoI/statistics are re-extracted from the stored
    result SELAFIN; runtime and sizes come from ``level.json``."""
    try:
        meta = json.loads((Path(level_dir) / "level.json").read_text())
        res = Path(level_dir) / "model" / getattr(cfg, "results_slf", "r2d.slf")
        data = selafin.read_slf(res, frame=-1)
        min_edge, mean_depth, mean_vel = _geom_stats(data, probes)
        dt = _cfl_dt(min_edge, mean_depth, mean_vel)
        qoi = extract_at_points(res, quantities, probes)
        cell_size = float(meta.get("cell_size") or meta["channel_size"])
        validity = None
        try:
            from hydromate import mesh_validity

            validity = mesh_validity.check_level(
                cfg, dx=cell_size, depth=mean_depth, velocity=mean_vel, ks=ks,
                pinned_model=pinned_model)
        except Exception as exc:                 # dummy test configs
            log.debug("validity check skipped for reused '%s': %s", label, exc)
        return LevelResult(label=label, cell_size=cell_size,
                           n_elem=int(meta.get("n_elem") or data["ikle"].shape[0]),
                           n_nodes=int(meta.get("n_nodes") or data["x"].size),
                           min_edge=min_edge, dt=dt,
                           runtime_s=meta.get("runtime_s"), qoi=qoi,
                           scale=float(meta.get("scale") or 1.0),
                           mean_depth=mean_depth, mean_velocity=mean_vel,
                           validity=validity)
    except Exception as exc:
        log.warning("could not reuse completed level '%s' (%s); re-running it",
                    label, exc)
        return None


# --------------------------------------------------------------------------- #
# extension support: runtime extrapolation + interactive prompt
# --------------------------------------------------------------------------- #


def estimate_next_runtime(levels: list[LevelResult],
                          next_cell_size: float) -> tuple[float | None, str]:
    """Estimated wall-clock [s] of a run at *next_cell_size*, with its basis.

    Fits ``log(runtime) = a + b log(cell_size)`` over the completed levels; when
    fewer than two distinct sizes carry runtimes, or the fitted exponent is
    implausible, falls back to the theoretical ``t ~ dx**-3`` scaling from the
    finest level (elements ~ dx**-2, CFL-bound steps ~ dx**-1). ``(None, ...)``
    when no runtimes were recorded (e.g. an injected test simulator).
    """
    pts = [(x.cell_size, x.runtime_s) for x in levels
           if x.runtime_s and x.cell_size and x.cell_size > 0 and x.runtime_s > 0]
    if not pts or next_cell_size <= 0:
        return None, "no runtimes recorded"
    if len({round(dx, 12) for dx, _ in pts}) >= 2:
        b, a = np.polyfit(np.log([dx for dx, _ in pts]),
                          np.log([t for _, t in pts]), 1)
        if -6.0 < b < -1.0:                      # sane scaling exponent
            return float(math.exp(a + b * math.log(next_cell_size))), \
                f"fitted t ~ dx^{b:.1f}"
    dx_f, t_f = pts[-1]
    return float(t_f * (dx_f / next_cell_size) ** 3), "theoretical t ~ dx^-3"


def _default_ask(prompt: str, *, default: bool = False) -> bool:
    """Yes/no prompt on stdin; returns *default* in a non-interactive session
    (so a batch/nohup run never hangs: extension stops, resume reuses)."""
    import sys

    if not sys.stdin.isatty():
        log.info("non-interactive session: '%s...' -> %s",
                 prompt.splitlines()[0][:70], "yes" if default else "no")
        return default
    suffix = " [Y/n] " if default else " [y/N] "
    try:
        reply = input(prompt + suffix).strip().lower()
    except EOFError:                             # stdin closed mid-run
        return default
    return default if not reply else reply in ("y", "yes")


def run_mesh_convergence(cfg: Config, *, discharge: float, levels=None,
                         quantities=DEFAULT_QUANTITIES, probes=None,
                         tolerance: float = DEFAULT_TOLERANCE, simulate=None,
                         base_dir: Path | None = None, runtime=None,
                         n_processors: int | None = None,
                         extend_ratio: float | None = None,
                         max_extra_levels: int = 3,
                         auto_extend: bool = False,
                         ask=None) -> ConvergenceReport:
    """Run the convergence study and return the :class:`ConvergenceReport`.

    *levels* is a list of ``(label, channel_size, floodplain_size)`` coarse->fine
    (default: one doubling around the config sizes; prefer :func:`ratio_levels`).
    The sizes are interpreted *relative to the config baseline* - each level is
    built with ``mesh.size_scale = channel_size / cfg.mesh.channel_size`` so gpkg
    per-zone sizes scale along (see :func:`make_telemac_simulator`). *simulate*
    overrides the TELEMAC runner (used in tests); by default each mesh is built
    and run once at constant *discharge*. *probes* default to the ground-truth /
    centerline points. *tolerance* is the GCI / relative-change goal (default 5%,
    per Celik et al. 2008).

    **Resume**: completed levels found under *base_dir* (result SELAFIN +
    ``level.json`` with matching sizes) are offered for reuse - QoI re-extracted,
    only missing levels run. **Extension**: with *extend_ratio* set, a
    not-converged study offers up to *max_extra_levels* additional refinements,
    each preceded by a runtime estimate and the mesh-validity check of the
    candidate size. Both go through *ask(prompt, default=...) -> bool* (default:
    stdin prompt; non-interactive sessions take the stated default: resume = yes,
    extend = no). Inject a callable for scripted runs/tests.

    **Auto-refinement**: with *auto_extend* set, resume and every extension are
    auto-approved (no prompt, interactive or not), so the study keeps refining one
    grid at a time until it reaches the GCI/relative-change goal, exhausts
    *max_extra_levels*, or a level fails to mesh/run (the practical refinement
    limit) - then it writes the report from the completed levels.
    """
    levels = list(levels or default_levels(cfg))
    quantities = list(quantities)
    probes = np.asarray(probes if probes is not None else default_probes(cfg), float)
    if base_dir is None:
        base_dir = cfg.postprocessing_path("mesh-convergence")
    base_dir = Path(base_dir)
    if auto_extend:
        # auto-refinement: approve resume + every extension without prompting, so
        # the study marches finer until it converges / hits max_extra_levels / a
        # level fails. An explicit *ask* (e.g. a test stub) still takes precedence.
        ask = ask or (lambda *a, **k: True)
    ask = ask or _default_ask
    simulate = simulate or make_telemac_simulator(
        cfg, discharge, base_dir=base_dir, runtime=runtime, n_processors=n_processors)

    # study context for the validity checks (graceful for bare test configs)
    try:
        from hydromate import steering

        pinned = steering.select_turbulence_model(cfg)[0]
    except Exception:
        pinned = None
    try:
        from hydromate import mesh_validity

        ks_channel = mesh_validity.channel_ks(cfg)
    except Exception:
        ks_channel = None

    reuse: dict[str, Path] = {}
    if base_dir.exists():
        done = _scan_completed(cfg, base_dir, levels)
        if done:
            summary = ", ".join(f"'{lbl}'" for lbl in done)
            if ask(f"mesh convergence: found {len(done)} completed level(s) in "
                   f"{base_dir} ({summary}).\nResume and reuse them, re-running "
                   "only the missing levels? ('n' re-runs everything; the "
                   "mesh-convergence report files are rewritten either way)",
                   default=True):
                reuse = done

    results: list[LevelResult] = []
    for label, ch, fp in levels:
        res = None
        if label in reuse:
            res = _load_completed_level(cfg, reuse[label], label, probes, quantities,
                                        ks=ks_channel, pinned_model=pinned)
            if res is not None:
                log.info("mesh convergence: reused completed '%s' (%d elements, "
                         "runtime %.0f s)", label, res.n_elem, res.runtime_s or 0.0)
        if res is None:
            log.info("mesh convergence: running '%s' (channel %.3f m, floodplain "
                     "%.3f m)", label, ch, fp)
            res = simulate(label, ch, fp, probes, quantities)
            log.info("  '%s': %d elements, shortest edge %.4f m, dt~%.4f s",
                     label, res.n_elem, res.min_edge, res.dt if res.dt else float("nan"))
        results.append(res)

    report = build_report(results, quantities, tolerance, probes)

    # extension: offer one finer mesh at a time while not converged
    extras = 0
    while extend_ratio and not report.converged and extras < max_extra_levels:
        label, ch, fp = _next_finer(levels, extend_ratio)
        finest = results[-1]
        cand_dx = (finest.cell_size / extend_ratio if finest.cell_size > 0 else ch)
        est, basis = estimate_next_runtime(results, cand_dx)
        est_txt = (f"~{est / 3600.0:.1f} h ({est:.0f} s, {basis})"
                   if est is not None else "unknown (no runtimes recorded)")
        validity_txt = ""
        try:
            from hydromate import mesh_validity

            cand = mesh_validity.check_level(
                cfg, dx=cand_dx,
                depth=finest.mean_depth if finest.mean_depth else 1.0,
                velocity=finest.mean_velocity if finest.mean_velocity else 1.0,
                ks=ks_channel, pinned_model=pinned,
                estimated=finest.mean_depth is None or finest.mean_velocity is None)
            validity_txt = "\n".join("  " + ln for ln in cand.lines())
            if cand.dx_below_ks or cand.regime_change:
                validity_txt += ("\n  *** VALIDITY WARNING: refining past this "
                                 "point strains the roughness/turbulence closures "
                                 "(see above) ***")
        except Exception as exc:
            log.debug("candidate validity check skipped: %s", exc)
        worst = max((report.gci_fine.get(q) if report.verdict_basis == "gci"
                     else (report.rel_change[q][-1] if report.rel_change.get(q)
                           else None)) or 0.0
                    for q in quantities)
        prompt = (f"mesh convergence: NOT converged after {len(results)} meshes "
                  f"(worst {'GCI' if report.verdict_basis == 'gci' else 'rel. change'}"
                  f" {worst * 100:.1f}% vs. {tolerance * 100:.0f}% goal).\n"
                  f"Run one more refinement '{label}' (channel {ch:.3f} m)?\n"
                  f"  estimated runtime: {est_txt}\n"
                  f"  validity at the candidate size:\n{validity_txt}\n"
                  "Continue with this refinement ('n' stops and writes the report)?")
        if not ask(prompt, default=False):
            log.info("mesh convergence: stopping without further refinement "
                     "(user choice / non-interactive default); writing the report")
            break
        try:
            res = simulate(label, ch, fp, probes, quantities)
        except Exception as exc:
            # a finer level can outrun the mesher/solver (e.g. BAMG cannot honour
            # the metric even after its retries, or the run diverges): treat it as
            # the practical refinement limit -- stop extending and report from the
            # levels already completed rather than crashing the whole study.
            log.error("mesh convergence: refinement '%s' (channel %.3f m) failed to "
                      "build/run (%s: %s); stopping extension and writing the report "
                      "from the %d completed level(s)", label, ch,
                      type(exc).__name__, exc, len(results))
            break
        log.info("  '%s': %d elements, shortest edge %.4f m, dt~%.4f s",
                 label, res.n_elem, res.min_edge, res.dt if res.dt else float("nan"))
        results.append(res)
        levels.append((label, ch, fp))
        extras += 1
        report = build_report(results, quantities, tolerance, probes)

    report.case = getattr(cfg, "name", "")
    rec = report.recommendation()
    log.info("mesh convergence: %s at %.1f%% tolerance (%s basis); recommended "
             "cell size %.3f m",
             "CONVERGED" if report.converged else "NOT converged",
             tolerance * 100, report.verdict_basis, rec["cell_size"])
    return report
