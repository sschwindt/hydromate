"""FlowTracker2 -> calibration-target hydraulics tab.

Reads SonTek FlowTracker2 exports and writes one row per measurement point into
the ``hydraulics`` tab of a ``calibration-target-data.xlsx`` template, keyed by
each point's own ID (which then joins the DGPS point layer in
``user-sources/geodata/``, see :mod:`hydromate.targets`).

Per point it extracts the three velocity components and their turbulent RMS
fluctuations:

* ``u_x, u_y, u_z``  <- the reported mean components ``VelX/VelY/VelZ`` (aka
  ``v(x)/v(y)/v(z)``).
* ``u_x', u_y', u_z'`` (RMS fluctuations)  <- the sample **standard deviation**
  ``u std`` / ``Std Dev v'(x)`` / ``sigma_x``. **These are NOT the ``VxErr``
  columns:** ``VxErr`` is the FlowTracker *standard error of the mean* (the
  velocity measurement uncertainty), i.e. ~``std / sqrt(Npts)``. When only a
  plain ``.ft.sum`` summary is available (``VxErr`` + ``Npts`` but no std-dev
  column), the RMS is reconstructed as ``u' ~ VxErr * sqrt(Npts)``.

``TKE`` is taken from a measured column when present, else left to the
template's live formula ``0.5*(u_x'^2 + u_y'^2 + u_z'^2)``; the water depth is
the local total/final depth. ``U_h`` and ``U_h'`` recompute from the components.

Three export layouts are auto-detected by their headers:

1. **SonTek ``.ft.sum``** (``ID, MeasD, Npts, VelX/VelY/VelZ, VxErr/VyErr/VzErr,
   FinalD``): only stderr -> RMS reconstructed.
2. **TKE-stats export** (``VelX...`` plus ``u std/v std/w std``): real RMS.
3. **FT_TKE_Summary** (``Point, v(x)..., Std Dev v'(x)..., TKE [m2/s2]``): real
   RMS + measured TKE; multiple worksheets scanned.
"""

from __future__ import annotations

import logging
import re
import zipfile
from pathlib import Path

import numpy as np
import pandas as pd

from hydromate.ground_truth import read_xlsx_sheet
from hydromate.targets import _norm_header

log = logging.getLogger("hydromate")


# --------------------------------------------------------------------------- #
# Header classification (robust to the three FlowTracker layouts)
# --------------------------------------------------------------------------- #
def _norm(h: object) -> str:
    """Lower-case a header and drop trailing unit annotations.

    Strips ``[m/s]``-style brackets and known unit parentheses (``(m/s)``,
    ``(deg)``, ...) but keeps a component parenthesis like ``(x)`` so that
    ``v(x)``/``v(y)``/``v(z)`` stay distinct.
    """
    s = str(h).strip().lower()
    s = re.sub(r"\[[^\]]*\]", "", s)
    s = re.sub(
        r"\((?:m/s|m2/s2|m²/s²|m|deg|°c|db|-|°|s|"
        r"ellipsoidal[^)]*|wgs[^)]*|dgps[^)]*)\)",
        "", s)
    return re.sub(r"\s+", " ", s).strip()


# canonical field -> the normalised headers that map to it, in priority order
_FIELD_HEADERS: dict[str, tuple[str, ...]] = {
    "u": ("velx", "v(x)", "vx", "u average", "average v(x)"),
    "v": ("vely", "v(y)", "vy", "v average", "average v(y)"),
    "w": ("velz", "v(z)", "vz", "w average", "average v(z)"),
    "u_std": ("u std", "std dev v’(x)", "std dev v'(x)", "σx", "sigmax"),
    "v_std": ("v std", "std dev v’(y)", "std dev v'(y)", "σy", "sigmay"),
    "w_std": ("w std", "std dev v’(z)", "std dev v'(z)", "σz", "sigmaz"),
    "u_err": ("vxerr", "v_err(x)", "u stderr"),
    "v_err": ("vyerr", "v_err(y)", "v stderr"),
    "w_err": ("vzerr", "v_err(z)", "w stderr"),
    "npts": ("npts",),
    "tke": ("tke",),                       # measured 3D TKE in m2/s2 (exact)
    "meas_depth": ("measd", "meas. depth", "meas depth"),
    "depth": ("finald", "total depth"),
    "id": ("id", "point", "#", "st", "st (dgps point id)"),
}
_VEL_HEADERS = {h for f in ("u", "v", "w") for h in _FIELD_HEADERS[f]}


def _build_column_map(header_cells: list) -> dict[str, int]:
    """Map canonical fields to column indices for one header row."""
    normed = {i: _norm(c) for i, c in enumerate(header_cells)}
    colmap: dict[str, int] = {}
    for field, candidates in _FIELD_HEADERS.items():
        for cand in candidates:
            hit = next((i for i, n in normed.items() if n == cand), None)
            if hit is not None:
                colmap[field] = hit
                break
    return colmap


def _to_float(value: object) -> float:
    try:
        f = float(value)
    except (TypeError, ValueError):
        return np.nan
    return f


def _clean_id(value: object) -> object | None:
    """Integral floats -> int (so 3.0 joins '3'); text IDs kept verbatim."""
    if value is None:
        return None
    s = str(value).strip()
    try:
        f = float(s)
    except ValueError:
        return s or None
    if np.isnan(f):
        return None
    return int(f) if f == int(f) else f


# --------------------------------------------------------------------------- #
# Reading
# --------------------------------------------------------------------------- #
def _sheet_count(path: Path) -> int:
    with zipfile.ZipFile(path) as z:
        return sum(1 for n in z.namelist()
                   if re.fullmatch(r"xl/worksheets/sheet\d+\.xml", n))


def _read_sheet_points(raw: pd.DataFrame, reconstruct_rms: bool) -> pd.DataFrame | None:
    """Extract per-point velocities/fluctuations from one raw worksheet frame."""
    header_idx = next(
        (i for i in range(len(raw))
         if any(_norm(c) in _VEL_HEADERS for c in raw.iloc[i])),
        None)
    if header_idx is None:
        return None
    colmap = _build_column_map(list(raw.iloc[header_idx]))
    if not all(f in colmap for f in ("u", "v", "w")):
        return None
    # a genuine summary/point table always carries a per-point error, std-dev or
    # TKE column; this rejects raw per-sample time-series sheets (which expose
    # only u/v/w) that a multi-sheet workbook may also contain.
    if not any(f in colmap for f in ("u_err", "v_err", "w_err",
                                     "u_std", "v_std", "w_std", "tke")):
        return None

    # data starts after the header, skipping a units row (non-numeric velocity)
    start = header_idx + 1
    if start < len(raw) and np.isnan(_to_float(raw.iloc[start, colmap["u"]])):
        start += 1

    records: list[dict] = []
    for i in range(start, len(raw)):
        row = raw.iloc[i]
        u = _to_float(row[colmap["u"]])
        if np.isnan(u):
            continue                        # blank / trailing row
        rec: dict[str, object] = {"u": u,
                                  "v": _to_float(row[colmap["v"]]),
                                  "w": _to_float(row[colmap["w"]])}
        rid = _clean_id(row[colmap["id"]]) if "id" in colmap else None
        rec["ID"] = rid if rid is not None else len(records)
        for f in ("u_std", "v_std", "w_std", "u_err", "v_err", "w_err",
                  "npts", "tke", "meas_depth", "depth"):
            if f in colmap:
                rec[f] = _to_float(row[colmap[f]])
        records.append(rec)

    if not records:
        return None
    df = pd.DataFrame(records)

    # RMS fluctuations: prefer the measured std dev; else reconstruct from the
    # standard error of the mean and the sample count (u' ~ stderr * sqrt(Npts)).
    npts = df["npts"] if "npts" in df else pd.Series(np.nan, index=df.index)
    for comp in ("u", "v", "w"):
        std_col, err_col = f"{comp}_std", f"{comp}_err"
        std = df[std_col] if std_col in df else pd.Series(np.nan, index=df.index)
        if reconstruct_rms and err_col in df:
            reconstructed = df[err_col] * np.sqrt(npts)
            std = std.where(std.notna(), reconstructed)
        df[std_col] = std

    keep = ["ID", "u", "v", "w", "u_std", "v_std", "w_std", "tke", "depth"]
    for c in keep:
        if c not in df:
            df[c] = np.nan
    return df[keep]


def read_flowtracker(path: Path, *, sheet: str | int | None = None,
                     reconstruct_rms: bool = True) -> pd.DataFrame:
    """Read one FlowTracker2 export -> per-point velocity/fluctuation table.

    Returns columns ``ID, u, v, w, u_std, v_std, w_std, tke, depth`` (one row
    per measurement point). All worksheets that look like a FlowTracker table
    are scanned and concatenated unless ``sheet`` restricts to one (name or
    0-based index). With ``reconstruct_rms`` the RMS fluctuations are filled
    from ``VxErr * sqrt(Npts)`` where a direct std-dev column is absent.
    """
    path = Path(path)
    if sheet is not None:
        raw = read_xlsx_sheet(path, sheet=sheet)
        df = _read_sheet_points(raw, reconstruct_rms)
        if df is None:
            raise ValueError(f"no FlowTracker velocity table found in {path.name} "
                             f"sheet {sheet!r}")
        parts = [df]
    else:
        parts = []
        for i in range(_sheet_count(path)):
            df = _read_sheet_points(read_xlsx_sheet(path, sheet=i), reconstruct_rms)
            if df is not None:
                log.info("  %s sheet %d: %d FlowTracker point(s)", path.name, i, len(df))
                parts.append(df)
        if not parts:
            raise ValueError(f"no FlowTracker velocity table found in {path.name}")
    return pd.concat(parts, ignore_index=True)


def read_flowtrackers(paths, *, reconstruct_rms: bool = True) -> pd.DataFrame:
    """Read and concatenate several FlowTracker2 exports (e.g. day1 + day2)."""
    frames = [read_flowtracker(p, reconstruct_rms=reconstruct_rms) for p in paths]
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# --------------------------------------------------------------------------- #
# Writing into the calibration-target hydraulics tab
# --------------------------------------------------------------------------- #
def fill_template_hydraulics(template: Path, sources, *, sheet_name: str = "hydraulics",
                             start_row: int = 2, reconstruct_rms: bool = True) -> int:
    """Write FlowTracker points into the template's hydraulics tab, keyed by ID.

    ``sources`` is one path or a list of FlowTracker2 exports. Existing rows in
    the tab are cleared first (so re-running is idempotent). ``u_x/u_y/u_z`` and
    ``u_x'/u_y'/u_z'`` are written; ``U_h``, ``U_h'`` and (unless a measured TKE
    is present) ``TKE`` keep their live formulas; the water depth is filled.
    Returns the number of points written.
    """
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    template = Path(template)
    if not template.exists():
        raise FileNotFoundError(
            f"template not found: {template}. Run `hydromate targets <config>` first.")
    paths = [Path(sources)] if isinstance(sources, (str, Path)) else [Path(p) for p in sources]
    points = read_flowtrackers(paths, reconstruct_rms=reconstruct_rms)
    if points.empty:
        log.warning("no FlowTracker points found in %s", [p.name for p in paths])
        return 0

    wb = load_workbook(template)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"template {template.name} has no {sheet_name!r} sheet")
    ws = wb[sheet_name]

    # locate columns from the header row (robust to any future reordering)
    header = {(_norm_header(ws.cell(row=1, column=c).value)): c
              for c in range(1, ws.max_column + 1)}
    need = {"id": "id", "u": "u_x", "v": "u_y", "w": "u_z",
            "u_std": "u_x'", "v_std": "u_y'", "w_std": "u_z'",
            "u_h": "u_h", "u_h_std": "u_h'", "tke": "tke", "depth": "water depth"}
    missing = [h for h in need.values() if h not in header]
    if missing:
        raise ValueError(f"{template.name} {sheet_name!r} tab is missing columns {missing}")
    col = {k: header[v] for k, v in need.items()}
    letter = {k: get_column_letter(c) for k, c in col.items()}

    # clear any previously-written rows (data columns only; keep the formulas)
    data_cols = [col[k] for k in ("id", "u", "v", "w", "u_std", "v_std", "w_std", "depth")]
    last = start_row - 1
    for r in range(start_row, ws.max_row + 1):
        if ws.cell(row=r, column=col["id"]).value not in (None, ""):
            last = r
    for r in range(start_row, last + 1):
        for c in data_cols:
            ws.cell(row=r, column=c).value = None

    def _cell(r: int, key: str, value) -> None:
        if value is not None and not (isinstance(value, float) and np.isnan(value)):
            ws.cell(row=r, column=col[key]).value = round(float(value), 6) \
                if key != "id" else value

    for i, (_, p) in enumerate(points.iterrows()):
        r = start_row + i
        _cell(r, "id", p["ID"])
        for k in ("u", "v", "w", "u_std", "v_std", "w_std", "depth"):
            _cell(r, k, p[k])
        # (re)write the derived-quantity formulas so they are correct regardless
        # of the tab's prior state
        ws.cell(row=r, column=col["u_h"]).value = (
            f'=IF(OR({letter["u"]}{r}="",{letter["v"]}{r}=""),"",'
            f'SQRT({letter["u"]}{r}^2+{letter["v"]}{r}^2))')
        ws.cell(row=r, column=col["u_h_std"]).value = (
            f'=IF(OR({letter["u_std"]}{r}="",{letter["v_std"]}{r}=""),"",'
            f'SQRT({letter["u_std"]}{r}^2+{letter["v_std"]}{r}^2))')
        if pd.notna(p["tke"]):
            _cell(r, "tke", p["tke"])       # measured TKE wins over the proxy
        else:
            ws.cell(row=r, column=col["tke"]).value = (
                f'=IF(OR({letter["u_std"]}{r}="",{letter["v_std"]}{r}="",'
                f'{letter["w_std"]}{r}=""),"",0.5*({letter["u_std"]}{r}^2+'
                f'{letter["v_std"]}{r}^2+{letter["w_std"]}{r}^2))')

    wb.save(template)
    n_rms = int(points[["u_std", "v_std", "w_std"]].notna().all(axis=1).sum())
    log.info("wrote %d FlowTracker point(s) into %s (%r tab); %d with full RMS "
             "fluctuations", len(points), template.name, sheet_name, n_rms)
    return len(points)


# thin runnable driver dropped next to the generated template (see
# hydromate.targets.write_target_template) so users get a co-located script.
DRIVER_SCRIPT = '''#!/usr/bin/env python
"""Extract FlowTracker2 velocities + RMS fluctuations into the hydraulics tab of
calibration-target-data.xlsx (this folder), keyed by each point's ID.

Run in the hydromate env, from this folder:

    mamba run -n hydromate-env python extract_flowtracker.py FILE1.xlsx [FILE2.xlsx ...]

With no arguments it picks up *FlowTracker*/*ft*sum* .xlsx files in this folder.
Each point becomes one row keyed by its ID (must match the ID field of the
ground_truth.targets.hydraulics_positions layer). u_x/u_y/u_z come from
VelX/VelY/VelZ; the RMS fluctuations u_x'/u_y'/u_z' come from the sample standard
deviation (u std / Std Dev v'(x) / sigma) - NOT VxErr, which is the standard
error of the mean (reconstructed as VxErr*sqrt(Npts) only when no std-dev column
exists). U_h, U_h' and TKE recompute from these in the sheet.
"""
import sys
from pathlib import Path

from hydromate.flowtracker import fill_template_hydraulics

HERE = Path(__file__).resolve().parent
TEMPLATE = HERE / "calibration-target-data.xlsx"


def main(argv):
    files = [Path(a) for a in argv]
    if not files:
        files = sorted(HERE.glob("*[Ff]low[Tt]racker*.xlsx")) + sorted(HERE.glob("*ft*sum*"))
    if not files:
        print("usage: python extract_flowtracker.py <flowtracker2 xlsx> [more ...]")
        return 1
    n = fill_template_hydraulics(TEMPLATE, files)
    print(f"wrote {n} FlowTracker point(s) into {TEMPLATE.name}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
'''


def write_driver_script(path: Path, force: bool = False) -> Path | None:
    """Drop the ``extract_flowtracker.py`` driver next to the template."""
    path = Path(path)
    if path.exists() and not force:
        return None
    path.write_text(DRIVER_SCRIPT)
    return path
