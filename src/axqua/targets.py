"""Calibration-target template: a user-friendly ``calibration-target-data.xlsx``.

Ground-truth data for the Bayesian calibration needs a *consistent* structure:
measured values keyed by **unique IDs** that link to point layers (GeoPackage /
shapefile) in ``user-sources/geodata/``, so positions (surveyed once, in *some*
CRS) and values (typed into a spreadsheet) stay cleanly separated but joinable.

This module writes and reads that structure:

* :func:`write_target_template` generates the template workbook (default
  ``user-sources/ground-truth/calibration-target-data.xlsx``) with four tabs:

  - ``hydraulics`` - one row per measurement vertical: ``ID`` (or explicit
    ``x/y``), velocity components ``u_x, u_y, u_z``, the horizontal magnitude
    ``U_h`` (auto = sqrt(u_x^2 + u_y^2), live Excel formula), the RMS turbulent
    fluctuations ``u_x', u_y', u_z', U_h'``, ``TKE`` (auto proxy
    0.5*(u_x'^2 + u_y'^2 + u_z'^2) when not measured directly), water depth and
    bottom elevation.
  - ``morphodynamics`` - grain-size characteristics per sample point
    (``d16..d90`` [mm], fine fraction < 1 mm) plus a ``dz DoD`` column that is
    **auto-filled on ingest** from the DEM-of-Difference raster whenever the
    case provides one (``geodata.dem_target`` / ``geodata.dem_of_difference``).
  - ``parameters`` - the calibration parameters: a drop-down over the
    parameter catalog, a zone/class column, the current value, min/max test
    range, and a tips column (live lookup of the catalog's typical ranges).
    Rows for the case's friction zones (from ``geodata.roughness_zones`` +
    ``roughness_table``) are prefilled with their current ks.
  - ``parameter-catalog`` - the reference table of calibration-relevant
    TELEMAC-2D / TELEMAC-3D / GAIA keywords with default ranges and tips;
    feeds the drop-down and the tips lookup.

* :func:`read_targets` ingests the filled template into the tidy ground-truth
  tables (joining IDs to the configured position layers, reprojected to the
  project CRS; grain sizes converted mm -> m; DoD sampled at the sample points).
* :func:`read_target_parameters` turns the ``parameters`` tab into
  :class:`~axqua.config.CalibrationParameter` rows with HydroBayesCal names
  (``zone<N>``, ``gaia<KEYWORD> <class>``, or the literal TELEMAC keyword).

The template stays robust to Excel quirks: ingest never trusts the cached
formula results for derived quantities (``U_h``, ``U_h'``, ``TKE`` are
recomputed from the components when blank), and the workbook is read with the
zip-level reader from :mod:`axqua.ground_truth`.
"""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from pathlib import Path

import numpy as np
import pandas as pd

from axqua.config import CalibrationParameter, Config
from axqua.ground_truth import read_xlsx_sheet

log = logging.getLogger("axqua")

TEMPLATE_NAME = "calibration-target-data.xlsx"
_N_INPUT_ROWS = 200        # rows carrying live formulas / drop-downs below the header


# --------------------------------------------------------------------------- #
# Parameter catalog: calibration-relevant TELEMAC-2D / TELEMAC-3D / GAIA keywords.
# Keyword spellings verified against the v9 dictionaries (telemac2d.dico,
# telemac3d.dico, gaia.dico). ``per_class`` marks parameters needing a zone /
# sediment-class / layer number next to them.
# --------------------------------------------------------------------------- #
@dataclass(frozen=True)
class ParameterSpec:
    label: str          # drop-down text shown to the user
    module: str         # friction | telemac2d | telemac3d | gaia
    keyword: str        # literal TELEMAC keyword ("" for the friction-zone ks)
    per_class: str      # "" or what the Zone/Class column means (zone/class/layer)
    vmin: float
    vmax: float
    tips: str

    def hbc_name(self, klass: int | None) -> str:
        """HydroBayesCal parameter name for this spec (and zone/class number)."""
        if self.module == "friction":
            if klass is None:
                raise ValueError(f"{self.label!r} needs a zone id in the Zone/Class column")
            return f"zone{klass}"
        if self.module == "gaia":
            name = f"gaia{self.keyword}"
            return f"{name} {klass}" if klass is not None else name
        if self.module == "gainlose":
            # HydroBayesCal rewrites a line in the FORTRAN FILE for any parameter
            # prefixed "f." - that is how a constant living in the generated
            # USER_RAIN routine (rather than in a .cas keyword) is calibrated.
            return f"f.{self.keyword}"
        return self.keyword


PARAMETER_CATALOG: tuple[ParameterSpec, ...] = (
    # --- friction ---------------------------------------------------------- #
    ParameterSpec(
        "FRICTION ZONE ks (Nikuradse)", "friction", "", "zone", 0.05, 1.0,
        "Nikuradse ks [m] of one roughness zone (Zone ID from roughness-zones.gpkg); "
        "gravel bed ~2-3 x d90 (channel 0.05-0.5), vegetated floodplain 0.1-1.0"),
    ParameterSpec(
        "ROUGHNESS COEFFICIENT OF BOUNDARIES", "telemac2d",
        "ROUGHNESS COEFFICIENT OF BOUNDARIES", "", 20.0, 80.0,
        "lateral-wall friction (unit set by LAW OF FRICTION ON LATERAL BOUNDARIES; "
        "Strickler Kst here): smooth walls 70-80, natural banks 20-35"),
    # --- turbulent viscosity / diffusivity --------------------------------- #
    ParameterSpec(
        "VELOCITY DIFFUSIVITY", "telemac2d", "VELOCITY DIFFUSIVITY", "", 1e-4, 1e-1,
        "turbulent eddy viscosity [m2/s]; only governs mixing with the "
        "constant-viscosity TURBULENCE MODEL 1 (rivers typically 0.01-0.1); too "
        "high smears recirculation zones and velocity gradients"),
    ParameterSpec(
        "COEFFICIENT FOR DIFFUSION OF TRACERS", "telemac2d",
        "COEFFICIENT FOR DIFFUSION OF TRACERS", "", 1e-4, 1e-1,
        "tracer eddy diffusivity [m2/s] (suspended load, heat, salinity); often "
        "set equal to the velocity diffusivity"),
    ParameterSpec(
        "NON-DIMENSIONAL DISPERSION COEFFICIENTS", "telemac2d",
        "NON-DIMENSIONAL DISPERSION COEFFICIENTS", "", 0.6, 6.0,
        "Elder model (TURBULENCE MODEL 2): longitudinal ~6.0, transverse ~0.6; "
        "two values in the .cas"),
    ParameterSpec(
        "PRODUCTION COEFFICIENT FOR SECONDARY CURRENTS", "telemac2d",
        "PRODUCTION COEFFICIENT FOR SECONDARY CURRENTS", "", 1.0, 10.0,
        "transversal (bend) currents, needs SECONDARY CURRENTS: YES; default "
        "7.071; raise to strengthen the outer-bank velocity shift in bends"),
    ParameterSpec(
        "DISSIPATION COEFFICIENT FOR SECONDARY CURRENTS", "telemac2d",
        "DISSIPATION COEFFICIENT FOR SECONDARY CURRENTS", "", 2.0, 10.0,
        "decay of the secondary-current intensity downstream of a bend; default 5.0"),
    # --- minimum depths ----------------------------------------------------- #
    ParameterSpec(
        "MINIMUM VALUE OF DEPTH", "telemac2d", "MINIMUM VALUE OF DEPTH", "",
        1e-3, 5e-2,
        "wetting/drying depth floor [m] (used with H CLIPPING); mesh-size "
        "dependent; too high dries shallow margins that carry ground truth"),
    ParameterSpec(
        "MINIMAL VALUE FOR DEPTH", "telemac3d", "MINIMAL VALUE FOR DEPTH", "",
        1e-3, 5e-2,
        "TELEMAC-3D dry-column guard for the sigma transform [m]"),
    # --- TELEMAC-3D diffusion ----------------------------------------------- #
    ParameterSpec(
        "COEFFICIENT FOR HORIZONTAL DIFFUSION OF VELOCITIES", "telemac3d",
        "COEFFICIENT FOR HORIZONTAL DIFFUSION OF VELOCITIES", "", 1e-4, 1e-1,
        "[m2/s]; only governs mixing with a constant-viscosity horizontal model"),
    ParameterSpec(
        "COEFFICIENT FOR VERTICAL DIFFUSION OF VELOCITIES", "telemac3d",
        "COEFFICIENT FOR VERTICAL DIFFUSION OF VELOCITIES", "", 1e-6, 1e-3,
        "[m2/s]; vertical mixing is much weaker than horizontal (depth-limited "
        "eddies); with a mixing-length/k-epsilon vertical model this is unused"),
    ParameterSpec(
        "COEFFICIENT FOR HORIZONTAL DIFFUSION OF TRACERS", "telemac3d",
        "COEFFICIENT FOR HORIZONTAL DIFFUSION OF TRACERS", "", 1e-4, 1e-1,
        "[m2/s] tracer/suspended-load horizontal diffusivity"),
    ParameterSpec(
        "COEFFICIENT FOR VERTICAL DIFFUSION OF TRACERS", "telemac3d",
        "COEFFICIENT FOR VERTICAL DIFFUSION OF TRACERS", "", 1e-6, 1e-3,
        "[m2/s] tracer/suspended-load vertical diffusivity"),
    ParameterSpec(
        "PRANDTL NUMBER", "telemac3d", "PRANDTL NUMBER", "", 0.5, 2.0,
        "turbulent Prandtl/Schmidt number (eddy viscosity / eddy diffusivity); "
        "~0.7-1.0 in free shear flows"),
    # --- gain-lose reach (porous body) --------------------------------------- #
    ParameterSpec(
        "POROUS ZONE kf (gain-lose)", "gainlose", "HMP_KF", "", 1e-6, 1e-2,
        "saturated hydraulic conductivity [m/s] of the porous body - the parameter "
        "that sets how much water a gain-lose reach exchanges. Riverbed values span "
        "orders of magnitude, so calibrate rather than assume: clean gravel "
        "1e-2..1e-1, moderately colmated 1e-4..1e-3, silted 1e-7..1e-5 (Calver 2001, "
        "doi:10.1111/j.1745-6584.2001.tb02343.x, pools published riverbed data over "
        "1e-9..1e-2 m/s). Needs gain_lose.conductivity set (not a prescribed "
        "discharge, which fixes the exchange and ignores kf)"),
    # --- GAIA (morphodynamics) ---------------------------------------------- #
    ParameterSpec(
        "CLASSES SHIELDS PARAMETERS", "gaia", "CLASSES SHIELDS PARAMETERS",
        "class", 0.03, 0.06,
        "critical Shields stress [-] per sediment class; gravel 0.03-0.06 "
        "(0.047 classic Meyer-Peter & Mueller); THE primary bedload knob"),
    ParameterSpec(
        "CLASSES SEDIMENT DIAMETERS", "gaia", "CLASSES SEDIMENT DIAMETERS",
        "class", 1e-4, 0.2,
        "characteristic diameter [m] per class; vary within the measured GSD "
        "uncertainty (e.g. around d50 from the morphodynamics tab)"),
    ParameterSpec(
        "MPM COEFFICIENT", "gaia", "MPM COEFFICIENT", "", 3.97, 13.0,
        "Meyer-Peter & Mueller transport coefficient: 8 classic, 3.97 "
        "Wong & Parker (2006), up to ~13 for steep gravel-bed rivers"),
    ParameterSpec(
        "RATIO BETWEEN SKIN FRICTION AND MEAN DIAMETER", "gaia",
        "RATIO BETWEEN SKIN FRICTION AND MEAN DIAMETER", "", 1.0, 6.0,
        "skin-friction roughness ks' = ratio x d50 (with SKIN FRICTION "
        "CORRECTION 1); ~3 typical; controls the effective transport stress"),
    ParameterSpec(
        "CLASSES HIDING FACTOR", "gaia", "CLASSES HIDING FACTOR", "class",
        0.4, 1.0,
        "per-class hiding/exposure factor [-] (HIDING FACTOR FORMULA 0); <1 "
        "shelters fine classes in a graded bed"),
    ParameterSpec(
        "ACTIVE LAYER THICKNESS", "gaia", "ACTIVE LAYER THICKNESS", "",
        0.01, 0.5,
        "[m]; ~1-3 x d90 (Hirano); limits sediment available for entrainment "
        "per time step and damps or frees the computed bed evolution"),
    ParameterSpec(
        "SECONDARY CURRENTS ALPHA COEFFICIENT", "gaia",
        "SECONDARY CURRENTS ALPHA COEFFICIENT", "", 0.2, 1.5,
        "bend effect on the bedload direction (with SECONDARY CURRENTS); "
        "default 1.0; lower on coarse beds"),
    ParameterSpec(
        "LAYERS NON COHESIVE BED POROSITY", "gaia",
        "LAYERS NON COHESIVE BED POROSITY", "layer", 0.25, 0.45,
        "bed porosity [-] per layer (sand/gravel ~0.30-0.40); scales transport "
        "into bed-level change, so it shifts DoD-matched evolution directly"),
    ParameterSpec(
        "CLASSES SETTLING VELOCITIES", "gaia", "CLASSES SETTLING VELOCITIES",
        "class", 1e-4, 1e-1,
        "[m/s] per suspended class; -9 lets GAIA compute it (van Rijn); "
        "calibrate within the range implied by the class diameter"),
    ParameterSpec(
        "LAYERS PARTHENIADES CONSTANT", "gaia", "LAYERS PARTHENIADES CONSTANT",
        "layer", 1e-5, 1e-3,
        "[kg/m2/s] mud erosion-rate constant (cohesive beds only)"),
    ParameterSpec(
        "LAYERS CRITICAL EROSION SHEAR STRESS OF THE MUD", "gaia",
        "LAYERS CRITICAL EROSION SHEAR STRESS OF THE MUD", "layer", 0.1, 10.0,
        "[Pa] per bed layer (cohesive); consolidated mud is at the high end"),
    ParameterSpec(
        "CLASSES CRITICAL SHEAR STRESS FOR MUD DEPOSITION", "gaia",
        "CLASSES CRITICAL SHEAR STRESS FOR MUD DEPOSITION", "class", 0.05, 1000.0,
        "[Pa]; deposition happens below this stress; 1000 = always deposits"),
    ParameterSpec(
        "MORPHOLOGICAL FACTOR", "gaia", "MORPHOLOGICAL FACTOR", "", 1.0, 100.0,
        "time acceleration, NOT a physics parameter - keep 1 during calibration "
        "unless run time forces otherwise (then validate the factor separately)"),
    ParameterSpec(
        "BED-LOAD TRANSPORT FORMULA FOR ALL SANDS", "gaia",
        "BED-LOAD TRANSPORT FORMULA FOR ALL SANDS", "", 1.0, 10.0,
        "CATEGORICAL (1=MPM, 4=Einstein, 5=Bijker, 6=Soulsby-van Rijn, 7=van "
        "Rijn, 30=MPM multi-class...) - compare discrete runs, do not sample "
        "continuously in the Bayesian calibration"),
)


def catalog_by_label() -> dict[str, ParameterSpec]:
    return {spec.label.lower(): spec for spec in PARAMETER_CATALOG}


# --------------------------------------------------------------------------- #
# Template layout
# --------------------------------------------------------------------------- #
HYDRAULICS_HEADERS = (
    "ID", "x [m]", "y [m]",
    "u_x [m/s]", "u_y [m/s]", "u_z [m/s]", "U_h [m/s]",
    "u_x' [m/s]", "u_y' [m/s]", "u_z' [m/s]", "U_h' [m/s]",
    "TKE [m2/s2]",
    "water depth [m]", "bottom elevation [m]",
)
MORPHO_HEADERS = (
    "ID", "x [m]", "y [m]",
    "d16 [mm]", "d30 [mm]", "d50 [mm]", "d60 [mm]", "d84 [mm]", "d90 [mm]",
    "fine fraction <1mm [-]",
    "dz DoD [m] (auto)",
)
PARAMETER_HEADERS = (
    "Parameter", "Zone / Class", "Current value", "Min", "Max",
    "Tips (typical range)",
)

# template header (unit brackets / "(auto)" stripped, lower-cased) -> tidy column
_HYDRAULICS_COLUMNS = {
    "id": "ID", "x": "x", "y": "y",
    "u_x": "u", "u_y": "v", "u_z": "w", "u_h": "u_mag",
    "u_x'": "u_std", "u_y'": "v_std", "u_z'": "w_std", "u_h'": "u_mag_std",
    "tke": "tke", "water depth": "h", "bottom elevation": "z",
}
_MORPHO_COLUMNS = {
    "id": "ID", "x": "x", "y": "y",
    "d16": "d16", "d30": "d30", "d50": "d50", "d60": "d60",
    "d84": "d84", "d90": "d90",
    "fine fraction <1mm": "fine_fraction", "dz dod": "dz",
}
_GRAIN_COLUMNS = ("d16", "d30", "d50", "d60", "d84", "d90")  # template mm -> tidy m


def _dod_status(cfg: Config) -> tuple[Path | None, str]:
    """Resolve the DoD raster for this case and a human-readable status line."""
    if cfg.geodata.dem_of_difference is not None:
        return Path(cfg.geodata.dem_of_difference), (
            f"precomputed raster: {cfg.geodata.dem_of_difference}")
    produced = cfg.preprocessing_path(cfg.dem_of_difference.output)
    if produced.exists():
        return produced, f"computed by preprocessing stage 1: {produced}"
    if cfg.dem_of_difference.enabled or cfg.morphodynamics.enabled \
            or cfg.geodata.dem_target is not None:
        return None, (
            "auto: computed by preprocessing stage 1 from dem_initial/dem_target "
            f"-> {produced} (run preprocessing.py first, then re-ingest)")
    return None, ("not available - set geodata.dem_target (or "
                  "geodata.dem_of_difference) in case-config.yml to enable it")


# --------------------------------------------------------------------------- #
# Template generation
# --------------------------------------------------------------------------- #
def write_target_template(cfg: Config, path: Path | None = None,
                          force: bool = False) -> Path:
    """Write the ``calibration-target-data.xlsx`` template for this case.

    Prefills the ``parameters`` tab with the case's friction zones (current ks
    from ``geodata.roughness_table``) and, when morphodynamics is on, a Shields
    row per configured sediment class; wires the parameter drop-down and the
    live tips lookup to the ``parameter-catalog`` sheet; declares the DoD status
    on the ``morphodynamics`` tab. Refuses to overwrite an existing (possibly
    user-filled) file unless ``force``.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    if path is None:
        t = cfg.ground_truth.targets
        path = Path(t.file) if t is not None and t.file is not None \
            else Path(cfg.config_dir) / "user-sources" / "ground-truth" / TEMPLATE_NAME
    path = Path(path)
    if path.exists() and not force:
        raise FileExistsError(
            f"{path} already exists (it may contain filled-in measurements); "
            "pass force=True / --force to overwrite it")
    path.parent.mkdir(parents=True, exist_ok=True)

    dod_raster, dod_note = _dod_status(cfg)
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDEBF7")
    auto_fill = PatternFill("solid", fgColor="F2F2F2")

    wb = Workbook()

    # --- README ------------------------------------------------------------ #
    ws = wb.active
    ws.title = "README"
    readme = [
        f"Calibration target data for case '{cfg.name}' (generated by axqua)",
        "",
        "Every measurement row needs a UNIQUE ID that matches the 'ID' field of the",
        "point layer configured in case-config.yml (ground_truth.targets.*_positions,",
        "a gpkg/shp in user-sources/geodata/). Coordinates then come from that layer,",
        "reprojected to the project CRS (EPSG:%d). Alternatively fill x/y directly" % cfg.crs_epsg,
        "(explicit x/y win over the position layer).",
        "",
        "hydraulics tab: velocity components u_x/u_y/u_z and their RMS fluctuations",
        "u_x'/u_y'/u_z' per vertical. U_h, U_h' and TKE carry live formulas",
        "(U_h = sqrt(u_x^2+u_y^2); TKE proxy = 0.5*(u_x'^2+u_y'^2+u_z'^2));",
        "overwrite TKE with a measured value if you have one. axqua recomputes",
        "blank derived cells on ingest, so the formulas are a convenience only.",
        "",
        "morphodynamics tab: grain sizes in mm (converted to m on ingest) and the",
        "fine fraction < 1 mm as 0..1. The 'dz DoD' column is filled AUTOMATICALLY",
        "from the DEM of Difference at each sample point - leave it blank.",
        f"DEM of Difference status: {dod_note}",
        "",
        "parameters tab: pick a calibration parameter from the drop-down (see the",
        "parameter-catalog sheet for all options + typical ranges), give the zone /",
        "sediment-class number where applicable, the current model value, and the",
        "Min/Max test range for the Bayesian calibration. The Tips column shows the",
        "catalog guidance for the selected parameter.",
        "",
        "Then reference this file in case-config.yml under ground_truth.targets and",
        "re-run preprocessing.py (or the axqua build) to regenerate the",
        "calibration CSV and the HydroBayesCal config.",
    ]
    for i, line in enumerate(readme, start=1):
        ws.cell(row=i, column=1, value=line)
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 90

    # --- data tabs ---------------------------------------------------------- #
    def _sheet(name: str, headers: tuple[str, ...], comments: dict[str, str],
               auto_cols: tuple[str, ...] = ()) -> None:
        w = wb.create_sheet(name)
        for j, h in enumerate(headers, start=1):
            cell = w.cell(row=1, column=j, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            if h in comments:
                cell.comment = Comment(comments[h], "axqua", height=120, width=320)
            w.column_dimensions[get_column_letter(j)].width = max(12, len(h) + 2)
        w.freeze_panes = "A2"

        for h in auto_cols:                     # grey the auto-computed columns
            j = headers.index(h) + 1
            for r in range(2, _N_INPUT_ROWS + 2):
                w.cell(row=r, column=j).fill = auto_fill

    _sheet("hydraulics", HYDRAULICS_HEADERS, comments={
        "ID": "unique ID matching the 'ID' field of "
              "ground_truth.targets.hydraulics_positions (point gpkg/shp); "
              "or fill x/y directly",
        "x [m]": f"optional: explicit easting (EPSG:{cfg.crs_epsg}); "
                 "overrides the position layer",
        "U_h [m/s]": "auto: sqrt(u_x^2 + u_y^2) - horizontal velocity magnitude "
                     "compared to TELEMAC SCALAR VELOCITY",
        "U_h' [m/s]": "auto: sqrt(u_x'^2 + u_y'^2); used as the measurement "
                      "error of U_h in the calibration CSV",
        "TKE [m2/s2]": "auto proxy: 0.5*(u_x'^2 + u_y'^2 + u_z'^2); overwrite "
                       "with a measured TKE if available",
        "bottom elevation [m]": "bed elevation z_b [m a.s.l.]; becomes the "
                                "point's z; falls back to the position layer's "
                                "z field when blank",
    }, auto_cols=("U_h [m/s]", "U_h' [m/s]", "TKE [m2/s2]"))

    ws_h = wb["hydraulics"]
    for r in range(2, _N_INPUT_ROWS + 2):       # live formulas for the derived cells
        ws_h.cell(row=r, column=7,
                  value=f'=IF(OR(D{r}="",E{r}=""),"",SQRT(D{r}^2+E{r}^2))')
        ws_h.cell(row=r, column=11,
                  value=f'=IF(OR(H{r}="",I{r}=""),"",SQRT(H{r}^2+I{r}^2))')
        ws_h.cell(row=r, column=12,
                  value=f'=IF(OR(H{r}="",I{r}="",J{r}=""),"",'
                        f'0.5*(H{r}^2+I{r}^2+J{r}^2))')

    _sheet("morphodynamics", MORPHO_HEADERS, comments={
        "ID": "unique ID matching the 'ID' field of "
              "ground_truth.targets.sediment_positions (point gpkg/shp); "
              "or fill x/y directly",
        "d50 [mm]": "grain sizes in mm (converted to m on ingest); cumulative "
                    "GSD percentiles of the surface sample",
        "fine fraction <1mm [-]": "mass fraction of grains finer than 1 mm, 0..1",
        "dz DoD [m] (auto)": "LEAVE BLANK - sampled automatically from the DEM "
                             f"of Difference on ingest. Status: {dod_note}",
    }, auto_cols=("dz DoD [m] (auto)",))

    _sheet("parameters", PARAMETER_HEADERS, comments={
        "Parameter": "pick from the drop-down; the full list with guidance is "
                     "on the parameter-catalog sheet",
        "Zone / Class": "roughness Zone ID / GAIA sediment class / bed layer "
                        "number, where the parameter needs one",
        "Min": "lower bound of the calibration test range",
        "Max": "upper bound of the calibration test range",
    })

    # --- parameter catalog + drop-down / tips wiring ------------------------ #
    ws_c = wb.create_sheet("parameter-catalog")
    cat_headers = ("Parameter", "Module", "TELEMAC keyword / HBC name",
                   "Zone- or class-based", "Default min", "Default max", "Tips")
    for j, h in enumerate(cat_headers, start=1):
        cell = ws_c.cell(row=1, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
    for i, spec in enumerate(PARAMETER_CATALOG, start=2):
        hbc = "zone<N>" if spec.module == "friction" else (
            f"gaia{spec.keyword}" + (" <class>" if spec.per_class else "")
            if spec.module == "gaia" else spec.keyword)
        for j, v in enumerate((spec.label, spec.module, hbc,
                               spec.per_class or "-", spec.vmin, spec.vmax,
                               spec.tips), start=1):
            ws_c.cell(row=i, column=j, value=v)
    for col, width in zip("ABCDEFG", (46, 11, 46, 18, 12, 12, 100)):
        ws_c.column_dimensions[col].width = width
    ws_c.freeze_panes = "A2"

    ws_p = wb["parameters"]
    n_cat = len(PARAMETER_CATALOG) + 1
    dv = DataValidation(
        type="list", formula1=f"'parameter-catalog'!$A$2:$A${n_cat}",
        allow_blank=True, showDropDown=False,
        promptTitle="Calibration parameter",
        prompt="Pick a parameter; see the parameter-catalog sheet for guidance.")
    ws_p.add_data_validation(dv)
    dv.add(f"A2:A{_N_INPUT_ROWS + 1}")

    # prefill: one row per friction zone with its current ks
    row = 2
    for zone_id, ks in sorted(_roughness_zones(cfg).items()):
        spec = PARAMETER_CATALOG[0]
        ws_p.cell(row=row, column=1, value=spec.label)
        ws_p.cell(row=row, column=2, value=int(zone_id))
        ws_p.cell(row=row, column=3, value=float(ks))
        ws_p.cell(row=row, column=4, value=round(float(ks) / 4.0, 4))
        ws_p.cell(row=row, column=5, value=round(float(ks) * 4.0, 4))
        ws_p.cell(row=row, column=6,
                  value=f"current ks = {ks} m; {spec.tips}")
        row += 1
    if cfg.morphodynamics.enabled:
        shields = catalog_by_label()["classes shields parameters"]
        n_classes = max(1, len(getattr(cfg.morphodynamics, "sediment_classes", []) or []))
        for klass in range(1, n_classes + 1):
            ws_p.cell(row=row, column=1, value=shields.label)
            ws_p.cell(row=row, column=2, value=klass)
            ws_p.cell(row=row, column=4, value=shields.vmin)
            ws_p.cell(row=row, column=5, value=shields.vmax)
            ws_p.cell(row=row, column=6, value=shields.tips)
            row += 1
    for r in range(row, _N_INPUT_ROWS + 2):     # live tips for user-picked rows
        ws_p.cell(row=r, column=6,
                  value=f'=IFERROR(VLOOKUP($A{r},'
                        f"'parameter-catalog'!$A$2:$G${n_cat},7,FALSE),\"\")")
    ws_p.column_dimensions["A"].width = 46
    ws_p.column_dimensions["F"].width = 100

    wb.save(path)
    log.info("wrote calibration-target template -> %s "
             "(%d catalog parameters; DoD: %s)", path, len(PARAMETER_CATALOG), dod_note)

    # drop the FlowTracker2 extraction driver next to the template so the user
    # has a co-located, runnable script to fill the hydraulics tab.
    from axqua.flowtracker import write_driver_script

    script = write_driver_script(path.parent / "extract_flowtracker.py", force=force)
    if script is not None:
        log.info("wrote FlowTracker2 extraction script -> %s", script)
    return path


def _roughness_zones(cfg: Config) -> dict[int, float]:
    """{zone_id: current roughness} from the configured roughness table (if any)."""
    if cfg.geodata.roughness_table is None or not Path(cfg.geodata.roughness_table).exists():
        return {}
    from axqua.mesh import read_roughness_table

    return read_roughness_table(Path(cfg.geodata.roughness_table))


# --------------------------------------------------------------------------- #
# Ingestion: filled template -> tidy ground-truth tables
# --------------------------------------------------------------------------- #
def _norm_header(h: object) -> str:
    """'u_x [m/s]' -> 'u_x'; 'dz DoD [m] (auto)' -> 'dz dod'."""
    s = re.sub(r"\s*\[[^\]]*\]", "", str(h))
    s = re.sub(r"\s*\(auto\)", "", s, flags=re.I)
    return s.strip().lower()


def _read_tab(path: Path, sheet: str, columns: dict[str, str]) -> pd.DataFrame | None:
    """Read one template tab -> DataFrame with tidy column names (or None)."""
    try:
        raw = read_xlsx_sheet(path, sheet=sheet)
    except (ValueError, IndexError):
        return None
    if raw.empty:
        return None
    header_idx = next((i for i in range(len(raw))
                       if _norm_header(raw.iloc[i, 0]) == "id"), None)
    if header_idx is None:
        return None
    headers = [_norm_header(h) for h in raw.iloc[header_idx]]
    data = raw.iloc[header_idx + 1:].copy()
    data.columns = [columns.get(h, h) for h in headers]
    data = data.loc[:, [c for c in data.columns if c in columns.values()]]
    ids = data["ID"].astype(str).str.strip().str.lower()
    data = data[data["ID"].notna() & ~ids.isin(("", "none", "nan"))]
    for c in data.columns:
        if c != "ID":
            data[c] = pd.to_numeric(data[c], errors="coerce")
    data = data.dropna(how="all", subset=[c for c in data.columns if c != "ID"])
    return data.reset_index(drop=True) if len(data) else None


def _norm_id(value: object) -> str:
    """'1', 1, 1.0, ' P12 ' -> a canonical join string ('1' / 'P12').

    xlsx numeric cells always come back as floats, while the position layer
    usually carries integer IDs; both sides are normalised so 1.0 joins 1.
    """
    s = str(value).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (ValueError, OverflowError):
        pass
    return s


def _join_positions(df: pd.DataFrame, positions: Path | None, crs_epsg: int,
                    join_key: str, tab: str) -> pd.DataFrame:
    """Fill x/y (and z fallback) from the position layer; explicit x/y win."""
    ids = df["ID"].map(_norm_id)
    dupes = sorted(ids[ids.duplicated()].unique())
    if dupes:
        raise ValueError(f"{tab} tab: duplicate IDs {dupes} - IDs must be unique")

    have_xy = df.get("x", pd.Series(np.nan, index=df.index)).notna() \
        & df.get("y", pd.Series(np.nan, index=df.index)).notna()
    if "x" not in df:
        df["x"] = np.nan
    if "y" not in df:
        df["y"] = np.nan
    if bool(have_xy.all()):
        return df

    if positions is None:
        missing = sorted(ids[~have_xy])
        raise ValueError(
            f"{tab} tab: rows {missing} have no x/y and no position layer is "
            f"configured (ground_truth.targets.*_positions)")

    import geopandas as gpd

    pts = gpd.read_file(Path(positions))
    if pts.crs is not None and pts.crs.to_epsg() != crs_epsg:
        pts = pts.to_crs(epsg=crs_epsg)
    pcols = {c.lower(): c for c in pts.columns}
    key = pcols.get(join_key.lower())
    if key is None:
        raise ValueError(
            f"position layer {Path(positions).name!r} has no {join_key!r} column "
            f"(has {[c for c in pts.columns if c != 'geometry']})")
    pos = pd.DataFrame({
        "_id": pts[key].map(_norm_id),
        "_x": pts.geometry.x.to_numpy(),
        "_y": pts.geometry.y.to_numpy(),
    })
    zcol = pcols.get("z")
    pos["_z"] = pd.to_numeric(pts[zcol], errors="coerce") if zcol else np.nan
    pos = pos.drop_duplicates("_id")

    merged = df.assign(_id=ids).merge(pos, on="_id", how="left")
    unmatched = merged.loc[merged["_x"].isna() & ~have_xy.to_numpy(), "_id"]
    if len(unmatched):
        raise ValueError(
            f"{tab} tab: IDs {sorted(unmatched)} not found in "
            f"{Path(positions).name!r} (join key {join_key!r})")
    take = ~have_xy.to_numpy()
    merged.loc[take, "x"] = merged.loc[take, "_x"]
    merged.loc[take, "y"] = merged.loc[take, "_y"]
    if "z" in merged:
        fill_z = merged["z"].isna() & merged["_z"].notna()
        merged.loc[fill_z, "z"] = merged.loc[fill_z, "_z"]
    elif merged["_z"].notna().any():
        merged["z"] = merged["_z"]
    return merged.drop(columns=["_id", "_x", "_y", "_z"])


def _sample_dod(df: pd.DataFrame, raster: Path) -> pd.Series:
    """Sample the DoD raster at the sample points (NaN where nodata/off-grid)."""
    import rasterio

    with rasterio.open(raster) as src:
        vals = np.array([v[0] for v in src.sample(zip(df["x"], df["y"]))],
                        dtype=float)
        if src.nodata is not None:
            vals[np.isclose(vals, src.nodata)] = np.nan
    return pd.Series(vals, index=df.index)


def read_targets(cfg: Config) -> dict[str, pd.DataFrame]:
    """Ingest the filled template -> tidy tables ``{category: DataFrame}``.

    * ``hydraulics``: ``x, y, z`` then ``u, v, w, u_mag, u_std, v_std, w_std,
      u_mag_std, tke, h`` - blank derived cells (``u_mag``, ``u_mag_std``,
      ``tke``) are recomputed from the components.
    * ``morphodynamics``: ``x, y, z`` then ``d16..d90`` (converted mm -> m),
      ``fine_fraction`` and ``dz`` (auto-sampled from the DoD raster).
    """
    t = cfg.ground_truth.targets
    if t is None or t.file is None:
        return {}
    path = Path(t.file)
    if not path.exists():
        raise FileNotFoundError(f"ground_truth.targets.file not found: {path}")

    out: dict[str, pd.DataFrame] = {}

    hyd = _read_tab(path, "hydraulics", _HYDRAULICS_COLUMNS)
    if hyd is not None:
        hyd = _join_positions(hyd, t.hydraulics_positions, cfg.crs_epsg,
                              t.join_key, "hydraulics")
        if {"u", "v"} <= set(hyd.columns):
            mag = np.sqrt(hyd["u"].astype(float) ** 2 + hyd["v"].astype(float) ** 2)
            hyd["u_mag"] = hyd["u_mag"].fillna(mag) if "u_mag" in hyd else mag
        if {"u_std", "v_std"} <= set(hyd.columns):
            mag = np.sqrt(hyd["u_std"].astype(float) ** 2
                          + hyd["v_std"].astype(float) ** 2)
            hyd["u_mag_std"] = hyd["u_mag_std"].fillna(mag) if "u_mag_std" in hyd else mag
        if {"u_std", "v_std", "w_std"} <= set(hyd.columns):
            tke = 0.5 * (hyd["u_std"].astype(float) ** 2
                         + hyd["v_std"].astype(float) ** 2
                         + hyd["w_std"].astype(float) ** 2)
            hyd["tke"] = hyd["tke"].fillna(tke) if "tke" in hyd else tke
        out["hydraulics"] = _finalise(hyd)

    mor = _read_tab(path, "morphodynamics", _MORPHO_COLUMNS)
    if mor is not None:
        mor = _join_positions(mor, t.sediment_positions, cfg.crs_epsg,
                              t.join_key, "morphodynamics")
        for c in _GRAIN_COLUMNS:
            if c in mor:
                mor[c] = mor[c].astype(float) / 1000.0     # mm -> m
        raster, note = _dod_status(cfg)
        if raster is not None and Path(raster).exists():
            dod = _sample_dod(mor, Path(raster))
            mor["dz"] = mor["dz"].fillna(dod) if "dz" in mor else dod
            log.info("morphodynamics targets: sampled DoD (%s) at %d points "
                     "(%d inside detected change)", Path(raster).name, len(mor),
                     int(mor["dz"].notna().sum()))
        else:
            log.info("morphodynamics targets: no DoD raster yet (%s)", note)
        out["morphodynamics"] = _finalise(mor)

    log.info("calibration targets ingested from %s: %s", path.name,
             {k: len(v) for k, v in out.items()} or "no filled tabs")
    return out


def _finalise(df: pd.DataFrame) -> pd.DataFrame:
    """Order x,y,z first, keep ID for traceability, drop all-empty columns."""
    if "z" not in df:
        df["z"] = 0.0
    df["z"] = pd.to_numeric(df["z"], errors="coerce").fillna(0.0)
    lead = ["x", "y", "z"]
    rest = [c for c in df.columns if c not in (*lead, "ID")]
    rest = [c for c in rest if df[c].notna().any()]
    cols = [*lead, *rest, *(["ID"] if "ID" in df else [])]
    return df[cols].reset_index(drop=True)


# --------------------------------------------------------------------------- #
# Ingestion: parameters tab -> CalibrationParameter list
# --------------------------------------------------------------------------- #
def read_target_parameters(cfg: Config) -> list[CalibrationParameter]:
    """Read the ``parameters`` tab into HydroBayesCal calibration parameters.

    Rows need a parameter (drop-down label, else a literal TELEMAC keyword) and
    a numeric Min < Max; incomplete rows are skipped with a warning. Names
    follow HydroBayesCal's convention (``zone<N>`` for friction zones,
    ``gaia<KEYWORD> <class>`` for GAIA, the literal keyword otherwise).
    """
    t = cfg.ground_truth.targets
    if t is None or t.file is None or not Path(t.file).exists():
        return []
    raw = _read_parameters_tab(Path(t.file))
    if raw is None:
        return []

    catalog = catalog_by_label()
    params: list[CalibrationParameter] = []
    for _, row in raw.iterrows():
        label = str(row["parameter"]).strip()
        if not label or label.lower() == "nan":
            continue
        vmin = pd.to_numeric(row.get("min"), errors="coerce")
        vmax = pd.to_numeric(row.get("max"), errors="coerce")
        if pd.isna(vmin) or pd.isna(vmax) or not vmin < vmax:
            log.warning("parameters tab: skipping %r (needs numeric Min < Max, "
                        "got %s..%s)", label, row.get("min"), row.get("max"))
            continue
        klass = pd.to_numeric(row.get("zone"), errors="coerce")
        klass = int(klass) if pd.notna(klass) else None
        spec = catalog.get(label.lower())
        try:
            if spec is not None:
                name = spec.hbc_name(klass)
                comment = spec.tips
            else:                                  # literal TELEMAC keyword
                name = f"{label} {klass}" if klass is not None else label
                comment = ""
        except ValueError as exc:
            log.warning("parameters tab: skipping %r (%s)", label, exc)
            continue
        params.append(CalibrationParameter(name=name, min=float(vmin),
                                           max=float(vmax), comment=comment))
    if params:
        log.info("parameters tab: %d calibration parameter(s): %s",
                 len(params), [p.name for p in params])
    return params


def _read_parameters_tab(path: Path) -> pd.DataFrame | None:
    try:
        raw = read_xlsx_sheet(path, sheet="parameters")
    except (ValueError, IndexError):
        return None
    if raw.empty:
        return None
    header_idx = next((i for i in range(len(raw))
                       if _norm_header(raw.iloc[i, 0]) == "parameter"), None)
    if header_idx is None:
        return None
    names = {"parameter": "parameter", "zone / class": "zone", "min": "min",
             "max": "max", "current value": "current"}
    headers = [_norm_header(h) for h in raw.iloc[header_idx]]
    data = raw.iloc[header_idx + 1:].copy()
    data.columns = [names.get(h, h) for h in headers]
    data = data[data["parameter"].notna()]
    return data.reset_index(drop=True) if len(data) else None
