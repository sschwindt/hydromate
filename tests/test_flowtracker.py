"""FlowTracker2 -> hydraulics tab extraction.

Builds synthetic FlowTracker2 workbooks in the three real export layouts,
checks the velocity/fluctuation extraction (including the RMS reconstruction
from the standard error of the mean), then fills a generated
``calibration-target-data.xlsx`` and confirms the round trip through
:func:`axqua.targets.read_targets`.

Requires ``axqua-env`` (openpyxl/geopandas). No TELEMAC needed.

Run via pytest: mamba run -n axqua-env pytest tests/test_flowtracker.py
"""

from __future__ import annotations

import math
from pathlib import Path

import pytest

CRS = 25832
X0, Y0 = 700000.0, 5340000.0


# --------------------------------------------------------------------------- #
# synthetic FlowTracker2 workbooks
# --------------------------------------------------------------------------- #
def _write_ftsum(path: Path) -> None:
    """SonTek .ft.sum layout: ID/MeasD/Npts/VelX.../VxErr.../FinalD (stderr only)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["Desktop_Software", None, None, None, "DISCHARGE:", 30.6, "CMS"])
    ws.append(["ID", "MeasD", "Npts", "VelX", "VelY", "VelZ",
               "VxErr", "VyErr", "VzErr", "FinalD"])
    ws.append(["()", "(m)", "()", "(m/s)", "(m/s)", "(m/s)",
               "(m/s)", "(m/s)", "(m/s)", "(m)"])
    # id, measd, npts, vx, vy, vz, vxerr, vyerr, vzerr, finald
    ws.append([0, 0.12, 100, 0.60, 0.80, 0.10, 0.02, 0.03, 0.05, 0.50])
    ws.append([1, 0.24, 64, 1.00, 0.00, 0.05, 0.01, 0.02, 0.04, 0.80])
    wb.save(path)


def _write_tkestats(path: Path) -> None:
    """TKE-stats layout: header row 0 with VelX plus u std/v std/w std (real RMS)."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["#", "VelX (m/s)", "VelY (m/s)", "VelZ (m/s)",
               "VxErr (m/s)", "u std (m/s)", "v std (m/s)", "w std (m/s)",
               "MeasD (m)"])
    ws.append([0, 0.60, 0.80, 0.10, 0.02, 0.11, 0.09, 0.05, 0.30])
    ws.append([1, 1.00, 0.00, 0.05, 0.01, 0.07, 0.06, 0.04, 0.45])
    wb.save(path)


def _write_ft_tke_summary(path: Path) -> None:
    """FT_TKE_Summary layout: a summary sheet + a raw time-series sheet to skip."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "KB8"
    ws.append(["Campaign", None, None])
    ws.append([None, None, None])
    ws.append([None, None, None])
    ws.append(["Point", "File(s)", "Station", "East.", "North.",
               "v(x) [m/s]", "v(y) [m/s]", "v(z) [m/s]",
               "v_err(x) [m/s]", "Std Dev  v’(x) [m/s]", "Std Dev  v’(y) [m/s]",
               "Std Dev  v’(z) [m/s]", "TKE [m²/s²]", "Total Depth [m]"])
    ws.append(["P1", "f", 0, X0, Y0, 0.60, 0.80, 0.10, 0.02,
               0.11, 0.09, 0.05, 0.0134, 0.50])
    ws.append(["P2", "f", 0, X0, Y0, 1.00, 0.00, 0.05, 0.01,
               0.07, 0.06, 0.04, 0.0050, 0.80])
    raw = wb.create_sheet("KB8_Raw_1")            # raw per-sample -> must be skipped
    raw.append(["#", "Jahr", "v(x)", "v(y)", "v(z)"])
    for i in range(50):
        raw.append([i, 2025, 0.6 + 0.01 * i, 0.1, 0.0])
    wb.save(path)


# --------------------------------------------------------------------------- #
# reading
# --------------------------------------------------------------------------- #
def test_ftsum_reconstructs_rms_from_stderr(tmp_path):
    from axqua.flowtracker import read_flowtracker

    _write_ftsum(tmp_path / "day1.ft.sum.xlsx")
    df = read_flowtracker(tmp_path / "day1.ft.sum.xlsx")
    assert len(df) == 2
    assert list(df["ID"]) == [0, 1]
    assert df.loc[0, "u"] == pytest.approx(0.60)
    assert df.loc[0, "w"] == pytest.approx(0.10)
    # RMS reconstructed as VxErr * sqrt(Npts) (NOT the stderr itself)
    assert df.loc[0, "u_std"] == pytest.approx(0.02 * math.sqrt(100))   # 0.20
    assert df.loc[0, "w_std"] == pytest.approx(0.05 * math.sqrt(100))   # 0.50
    assert df.loc[1, "v_std"] == pytest.approx(0.02 * math.sqrt(64))    # 0.16
    assert df.loc[0, "depth"] == pytest.approx(0.50)
    assert df["tke"].isna().all()             # no measured TKE in a plain .ft.sum


def test_tkestats_uses_measured_std(tmp_path):
    from axqua.flowtracker import read_flowtracker

    _write_tkestats(tmp_path / "up.xlsx")
    df = read_flowtracker(tmp_path / "up.xlsx")
    assert len(df) == 2
    # the std-dev columns are used directly, not reconstructed from VxErr
    assert df.loc[0, "u_std"] == pytest.approx(0.11)
    assert df.loc[0, "v_std"] == pytest.approx(0.09)
    assert df.loc[1, "w_std"] == pytest.approx(0.04)


def test_multisheet_skips_raw_and_reads_measured_tke(tmp_path):
    from axqua.flowtracker import read_flowtracker

    _write_ft_tke_summary(tmp_path / "FT_TKE_Summary.xlsx")
    df = read_flowtracker(tmp_path / "FT_TKE_Summary.xlsx")
    assert list(df["ID"]) == ["P1", "P2"]     # raw 50-row sheet skipped
    assert df.loc[0, "u"] == pytest.approx(0.60)
    assert df.loc[0, "u_std"] == pytest.approx(0.11)
    assert df.loc[0, "tke"] == pytest.approx(0.0134)   # measured TKE kept
    assert df.loc[0, "depth"] == pytest.approx(0.50)


# --------------------------------------------------------------------------- #
# filling the template + round trip
# --------------------------------------------------------------------------- #
@pytest.fixture()
def case(tmp_path):
    from axqua.config import load_config

    geo = tmp_path / "user-sources" / "geodata"
    gt = tmp_path / "user-sources" / "ground-truth"
    geo.mkdir(parents=True)
    gt.mkdir(parents=True)
    (geo / "roughness-table.csv").write_text("zone_id,ks\n1,0.2\n")

    import geopandas as gpd
    from shapely.geometry import Point

    gpd.GeoDataFrame(
        {"ID": [0, 1]},
        geometry=[Point(X0, Y0), Point(X0 + 10, Y0 + 5)], crs=f"EPSG:{CRS}",
    ).to_file(geo / "dgps.gpkg", driver="GPKG")

    cfg_yaml = tmp_path / "case-config.yml"
    cfg_yaml.write_text(f"""
project: {{name: ft-test, crs_epsg: {CRS}}}
telemac: {{pysource: {geo / 'roughness-table.csv'}}}
geodata:
  dem_initial: user-sources/geodata/dem.tif
  boundary: user-sources/geodata/roi.gpkg
  roughness_table: user-sources/geodata/roughness-table.csv
boundaries: {{liquid_boundaries: user-sources/geodata/lb.gpkg}}
ground_truth:
  targets:
    file: user-sources/ground-truth/calibration-target-data.xlsx
    hydraulics_positions: user-sources/geodata/dgps.gpkg
""")
    return load_config(cfg_yaml)


def test_write_target_template_emits_driver_script(case):
    from axqua.targets import write_target_template

    out = write_target_template(case)
    script = out.parent / "extract_flowtracker.py"
    assert script.exists()
    assert "fill_template_hydraulics" in script.read_text()


def test_fill_template_and_round_trip(case, tmp_path):
    from axqua.flowtracker import fill_template_hydraulics
    from axqua.targets import read_targets, write_target_template

    template = write_target_template(case)
    _write_ftsum(tmp_path / "day1.xlsx")
    n = fill_template_hydraulics(template, tmp_path / "day1.xlsx")
    assert n == 2

    from openpyxl import load_workbook

    wb = load_workbook(template)
    ws = wb["hydraulics"]
    assert ws["A2"].value == 0 and ws["A3"].value == 1
    assert ws["D2"].value == pytest.approx(0.60)     # u_x
    assert ws["H2"].value == pytest.approx(0.20)     # u_x' (reconstructed RMS)
    assert ws["M2"].value == pytest.approx(0.50)     # water depth
    assert str(ws["G2"].value).startswith("=")       # U_h keeps its formula
    assert str(ws["L2"].value).startswith("=")       # TKE proxy formula (no measured)

    # ingest: IDs join the DGPS layer, derived quantities recompute
    tables = read_targets(case)
    hyd = tables["hydraulics"]
    assert len(hyd) == 2
    assert hyd.loc[0, "x"] == pytest.approx(X0, abs=0.01)
    assert hyd.loc[0, "u_mag"] == pytest.approx(1.0)             # sqrt(.6^2+.8^2)
    # RMS reconstructed as VxErr*sqrt(100): u'=0.20, v'=0.30, w'=0.50
    assert hyd.loc[0, "tke"] == pytest.approx(0.5 * (0.20**2 + 0.30**2 + 0.50**2))


def test_fill_is_idempotent(case, tmp_path):
    from axqua.flowtracker import fill_template_hydraulics
    from axqua.targets import write_target_template

    template = write_target_template(case)
    _write_ftsum(tmp_path / "day1.xlsx")
    fill_template_hydraulics(template, tmp_path / "day1.xlsx")
    fill_template_hydraulics(template, tmp_path / "day1.xlsx")   # re-run

    from openpyxl import load_workbook

    ws = load_workbook(template)["hydraulics"]
    assert ws["A2"].value == 0 and ws["A3"].value == 1
    assert ws["A4"].value in (None, "")          # no stale duplicated rows
