"""Calibration-target template: generation and ingest round trip.

Builds a tiny synthetic case (roughness zones + table, DGPS-style point layers,
a small DoD raster), generates ``calibration-target-data.xlsx``, fills it in
programmatically (as a user would in Excel), and checks:

* the template carries the four tabs, the parameter drop-down and the
  prefilled friction-zone rows (current ks from the roughness table);
* :func:`hydromate.targets.read_targets` joins the hydraulics/morphodynamics
  IDs to the point layers (reprojected to the project CRS), recomputes the
  derived quantities (U_h, U_h', TKE proxy), converts grain sizes mm -> m and
  samples the DoD raster into ``dz``;
* :func:`hydromate.targets.read_target_parameters` turns the parameters tab
  into HydroBayesCal names/ranges (``zone<N>``, ``gaia<KEYWORD> <class>``,
  literal keywords) and these override the config's parameters in the merge;
* unique-ID violations and unmatched IDs raise.

Requires ``hydromate-env`` (geopandas/rasterio/openpyxl). No TELEMAC needed.

Run via pytest: mamba run -n hydromate-env pytest tests/test_targets.py
"""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pytest

CRS = 25832
X0, Y0 = 700000.0, 5340000.0


# --------------------------------------------------------------------------- #
# fixtures
# --------------------------------------------------------------------------- #
def _write_points(path: Path, ids, epsg: int, with_z: bool = True) -> None:
    import geopandas as gpd
    from shapely.geometry import Point

    xs = [X0 + 10.0 * i for i in range(len(ids))]
    ys = [Y0 + 5.0 * i for i in range(len(ids))]
    gdf = gpd.GeoDataFrame(
        {"ID": ids, **({"z": [400.0 + i for i in range(len(ids))]} if with_z else {})},
        geometry=[Point(x, y) for x, y in zip(xs, ys)], crs=f"EPSG:{CRS}",
    )
    if epsg != CRS:
        gdf = gdf.to_crs(epsg=epsg)
    gdf.to_file(path, driver="GPKG")


def _write_dod(path: Path, value: float = -0.35) -> None:
    import rasterio
    from rasterio.transform import from_origin

    data = np.full((40, 40), value, dtype="float32")
    data[0, 0] = -9999.0
    with rasterio.open(
        path, "w", driver="GTiff", height=40, width=40, count=1,
        dtype="float32", crs=f"EPSG:{CRS}", nodata=-9999.0,
        transform=from_origin(X0 - 20.0, Y0 + 150.0, 5.0, 5.0),
    ) as dst:
        dst.write(data, 1)


@pytest.fixture()
def case(tmp_path: Path):
    """A minimal loaded Config with roughness zones, position layers and a DoD."""
    from hydromate.config import load_config

    geo = tmp_path / "user-sources" / "geodata"
    gt = tmp_path / "user-sources" / "ground-truth"
    geo.mkdir(parents=True)
    gt.mkdir(parents=True)

    (geo / "roughness-table.csv").write_text("zone_id,ks\n1,0.2\n2,0.5\n")
    # position layers deliberately NOT in the project CRS (reprojection on ingest)
    _write_points(geo / "dgps-hydraulics.gpkg", [1, 2, 3], epsg=25833)
    _write_points(geo / "sediment-samples.gpkg", [10, 11], epsg=32632, with_z=False)
    _write_dod(geo / "dod.tif")

    cfg_yaml = tmp_path / "case-config.yml"
    cfg_yaml.write_text(f"""
project:
  name: targets-test
  crs_epsg: {CRS}
telemac:
  pysource: {geo / 'roughness-table.csv'}    # not sourced in this test
geodata:
  dem_initial: user-sources/geodata/dem.tif  # dummy (existence not checked here)
  boundary: user-sources/geodata/roi.gpkg
  roughness_table: user-sources/geodata/roughness-table.csv
  dem_of_difference: user-sources/geodata/dod.tif
boundaries:
  liquid_boundaries: user-sources/geodata/lb.gpkg
morphodynamics:
  enabled: true
  sediment_classes: [{{}}]
ground_truth:
  targets:
    file: user-sources/ground-truth/calibration-target-data.xlsx
    hydraulics_positions: user-sources/geodata/dgps-hydraulics.gpkg
    sediment_positions: user-sources/geodata/sediment-samples.gpkg
calibration:
  parameters:
    - {{ name: zone1, min: 0.01, max: 0.10, comment: "config value, overridden" }}
    - {{ name: zone9, min: 0.10, max: 0.90, comment: "config-only zone" }}
""")
    return load_config(cfg_yaml)


def _fill_template(path: Path) -> None:
    """Fill the generated template the way a user would."""
    from openpyxl import load_workbook

    wb = load_workbook(path)
    ws = wb["hydraulics"]
    # ID, x, y, u_x, u_y, u_z, U_h, ux', uy', uz', U_h', TKE, h, z_b
    ws["A2"], ws["D2"], ws["E2"], ws["F2"] = 1, 0.6, 0.8, 0.1
    ws["H2"], ws["I2"], ws["J2"] = 0.03, 0.04, 0.12
    ws["M2"], ws["N2"] = 1.2, 401.5
    for col in "GKL":                       # blank the formula cells (no cache)
        ws[f"{col}2"] = None
    ws["A3"], ws["D3"], ws["E3"], ws["M3"] = 2, 1.0, 0.0, 0.8
    ws["L3"] = 0.005                        # measured TKE wins over the proxy
    for col in "GK":
        ws[f"{col}3"] = None
    # row with explicit coordinates instead of an ID-joined position
    ws["A4"], ws["B4"], ws["C4"] = 3, X0 + 999.0, Y0 + 999.0
    ws["D4"], ws["E4"], ws["M4"] = 0.5, 0.5, 0.5
    for col in "GKL":
        ws[f"{col}4"] = None
    for r in range(5, 202):                 # clear the remaining formula rows
        for col in "GKL":
            ws[f"{col}{r}"] = None

    ws = wb["morphodynamics"]
    ws["A2"], ws["D2"], ws["F2"], ws["I2"], ws["J2"] = 10, 8.0, 32.0, 96.0, 0.12
    ws["A3"], ws["F3"] = 11, 45.0

    ws = wb["parameters"]
    # find the first free row below the prefilled ones
    row = 2
    while ws.cell(row=row, column=1).value not in (None, ""):
        val = str(ws.cell(row=row, column=1).value)
        if val.startswith("="):
            break
        row += 1
    ws.cell(row=row, column=1, value="MINIMUM VALUE OF DEPTH")
    ws.cell(row=row, column=4, value=0.001)
    ws.cell(row=row, column=5, value=0.02)
    ws.cell(row=row + 1, column=1, value="SOME CUSTOM KEYWORD")
    ws.cell(row=row + 1, column=4, value=1.0)
    ws.cell(row=row + 1, column=5, value=2.0)
    ws.cell(row=row + 2, column=1, value="VELOCITY DIFFUSIVITY")  # missing max -> skipped
    ws.cell(row=row + 2, column=4, value=0.001)
    for r in range(row, 202):               # tips formulas would read as junk text
        ws.cell(row=r + 3, column=1, value=None)
        ws.cell(row=r, column=6, value=None)
    wb.save(path)


# --------------------------------------------------------------------------- #
# template generation
# --------------------------------------------------------------------------- #
def test_template_layout_and_prefill(case):
    from openpyxl import load_workbook

    from hydromate.targets import PARAMETER_CATALOG, write_target_template

    out = write_target_template(case)
    assert out == Path(case.ground_truth.targets.file)

    wb = load_workbook(out)
    assert {"README", "hydraulics", "morphodynamics", "parameters",
            "parameter-catalog"} <= set(wb.sheetnames)

    ws = wb["parameters"]
    # friction zones prefilled from the roughness table with their current ks
    assert ws["A2"].value == "FRICTION ZONE ks (Nikuradse)"
    assert (ws["B2"].value, ws["C2"].value) == (1, 0.2)
    assert (ws["B3"].value, ws["C3"].value) == (2, 0.5)
    assert ws["D2"].value == pytest.approx(0.05) and ws["E2"].value == pytest.approx(0.8)
    # morphodynamics enabled -> a Shields row per sediment class
    assert ws["A4"].value == "CLASSES SHIELDS PARAMETERS" and ws["B4"].value == 1
    # the drop-down spans the whole catalog
    dv = list(ws.data_validations.dataValidation)
    assert len(dv) == 1 and str(len(PARAMETER_CATALOG) + 1) in dv[0].formula1
    # catalog sheet lists every spec
    assert wb["parameter-catalog"].max_row == len(PARAMETER_CATALOG) + 1

    # refuses to clobber a (possibly filled) template unless forced
    with pytest.raises(FileExistsError):
        write_target_template(case)
    write_target_template(case, force=True)


# --------------------------------------------------------------------------- #
# ingest round trip
# --------------------------------------------------------------------------- #
def test_read_targets_round_trip(case):
    from hydromate.targets import read_targets, write_target_template

    _fill_template(write_target_template(case))
    tables = read_targets(case)
    assert set(tables) == {"hydraulics", "morphodynamics"}

    hyd = tables["hydraulics"]
    assert list(hyd.columns[:3]) == ["x", "y", "z"]
    assert len(hyd) == 3
    # ID 1: position joined from the 25833 layer, reprojected back to 25832
    assert hyd.loc[0, "x"] == pytest.approx(X0, abs=0.01)
    assert hyd.loc[0, "y"] == pytest.approx(Y0, abs=0.01)
    # derived quantities recomputed from the components
    assert hyd.loc[0, "u_mag"] == pytest.approx(1.0)              # sqrt(.6^2+.8^2)
    assert hyd.loc[0, "u_mag_std"] == pytest.approx(0.05)         # sqrt(.03^2+.04^2)
    assert hyd.loc[0, "tke"] == pytest.approx(0.5 * (0.03**2 + 0.04**2 + 0.12**2))
    assert hyd.loc[0, "z"] == pytest.approx(401.5)                # bottom elevation wins
    assert hyd.loc[1, "tke"] == pytest.approx(0.005)              # measured TKE kept
    assert hyd.loc[1, "z"] == pytest.approx(401.0)                # layer z fallback
    # explicit x/y win over the position layer
    assert hyd.loc[2, "x"] == pytest.approx(X0 + 999.0)

    mor = tables["morphodynamics"]
    assert len(mor) == 2
    assert mor.loc[0, "d16"] == pytest.approx(0.008)              # mm -> m
    assert mor.loc[0, "d90"] == pytest.approx(0.096)
    assert mor.loc[0, "fine_fraction"] == pytest.approx(0.12)
    # DoD sampled at both sample points
    assert mor.loc[0, "dz"] == pytest.approx(-0.35)
    assert mor.loc[1, "dz"] == pytest.approx(-0.35)


def test_compile_and_calibration_csv(case):
    from hydromate import calibration
    from hydromate.ground_truth import compile_ground_truth, read_tidy
    from hydromate.targets import write_target_template

    _fill_template(write_target_template(case))
    out = compile_ground_truth(case)
    assert out is not None and out.exists()
    tidy = read_tidy(out)
    assert {"hydraulics", "morphodynamics"} <= set(tidy)

    case.calibration.calibration_quantities = ["SCALAR VELOCITY", "WATER DEPTH"]
    Path(case.calibration_dir).mkdir(parents=True, exist_ok=True)  # pipeline normally does
    csv = calibration.build_calibration_csv(case)
    assert csv is not None
    import pandas as pd

    df = pd.read_csv(csv)
    assert list(df.columns[:4]) == ["id", "x", "y", "z"]
    assert df["SCALAR VELOCITY_DATA"].iloc[0] == pytest.approx(1.0)
    # U_h' is the measured error of U_h
    assert df["SCALAR VELOCITY_ERROR"].iloc[0] == pytest.approx(0.05)
    assert df["WATER DEPTH_DATA"].iloc[0] == pytest.approx(1.2)


def test_read_target_parameters_and_merge(case):
    from hydromate import calibration
    from hydromate.targets import read_target_parameters, write_target_template

    _fill_template(write_target_template(case))
    params = {p.name: p for p in read_target_parameters(case)}
    # prefixed names: friction zones, GAIA class param, literal keywords
    assert params["zone1"].min == pytest.approx(0.05)
    assert params["zone1"].max == pytest.approx(0.8)
    assert "zone2" in params
    assert params["gaiaCLASSES SHIELDS PARAMETERS 1"].min == pytest.approx(0.03)
    assert params["MINIMUM VALUE OF DEPTH"].max == pytest.approx(0.02)
    assert "SOME CUSTOM KEYWORD" in params
    assert "VELOCITY DIFFUSIVITY" not in params        # incomplete row skipped

    merged = {p.name: p for p in calibration.merged_parameters(case)}
    assert merged["zone1"].min == pytest.approx(0.05)  # template overrides config
    assert merged["zone9"].min == pytest.approx(0.10)  # config-only entry kept


def test_duplicate_and_unmatched_ids_raise(case):
    from openpyxl import load_workbook

    from hydromate.targets import read_targets, write_target_template

    path = write_target_template(case)
    _fill_template(path)
    wb = load_workbook(path)
    ws = wb["hydraulics"]
    ws["A3"] = 1                                        # duplicate of row 2
    wb.save(path)
    with pytest.raises(ValueError, match="unique"):
        read_targets(case)

    _fill_template(write_target_template(case, force=True))
    wb = load_workbook(path)
    wb["morphodynamics"]["A3"] = 99                     # not in the point layer
    wb.save(path)
    with pytest.raises(ValueError, match="99"):
        read_targets(case)
